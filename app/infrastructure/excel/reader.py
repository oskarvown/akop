"""Единая абстракция чтения `.xls` (xlrd/BIFF8) и `.xlsx` (openpyxl).

См. `docs/DATA_CONTRACT.md` §3, §4: парсер должен читать значения ячеек по
физической позиции колонки независимо от `hidden`/ширины, и не зависеть от
конкретного движка чтения. Этот модуль скрывает разницу API xlrd/openpyxl за
одной структурой `RawSheet`/`RawRow`, используемой всем остальным пайплайном
парсера (`header_parser.py`, `row_classifier.py`, `validator.py`).

Пустые колонки справа от последней непустой ячейки (артефакт Excel: ширина
колонки без данных) и полностью пустые строки в конце листа (артефакт
`max_row`) обрезаются — см. `_trim_trailing_empty_columns`. Это не ослабляет
fingerprint: валидатор по-прежнему требует ровно 17 колонок с данными (A–Q) и
строку «Итого» последней; колонка с реальными данными за Q или непустые строки
после «Итого» остаются и приводят к отклонению.
"""
from __future__ import annotations

import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any


class ExcelReadError(Exception):
    """Файл повреждён, нечитаем или не соответствует ожидаемому формату книги."""


class SheetNotFoundError(ExcelReadError):
    """Лист с ожидаемым именем отсутствует в книге (см. `docs/DATA_CONTRACT.md` §4.1)."""


@dataclass(frozen=True)
class RawRow:
    """Одна физическая строка листа, с иерархическим уровнем (Excel group outline)."""

    index: int  # 0-based позиция строки в листе
    outline_level: int
    values: tuple[Any, ...]  # длина == RawSheet.n_cols, по физической позиции колонки


@dataclass(frozen=True)
class RawSheet:
    sheet_name: str
    n_rows: int
    n_cols: int
    rows: tuple[RawRow, ...]
    column_hidden: tuple[bool, ...]
    column_width: tuple[float | None, ...]


def _cell_is_empty(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def _effective_column_count(rows: list[RawRow], reported_n_cols: int) -> int:
    """Правая граница по содержимому: 1 + индекс последней непустой ячейки.

    Если во всём листе нет ни одной непустой ячейки — возвращает 0.
    """
    rightmost = -1
    for row in rows:
        for col_idx, value in enumerate(row.values):
            if not _cell_is_empty(value):
                if col_idx > rightmost:
                    rightmost = col_idx
    if rightmost < 0:
        return 0
    return min(rightmost + 1, reported_n_cols)


def _row_is_empty(row: RawRow) -> bool:
    return all(_cell_is_empty(value) for value in row.values)


def _trim_trailing_empty_rows(rows: list[RawRow]) -> list[RawRow]:
    """Убирает полностью пустые строки в конце листа (артефакт Excel max_row)."""
    end = len(rows)
    while end > 0 and _row_is_empty(rows[end - 1]):
        end -= 1
    return rows[:end]


def _trim_trailing_empty_columns(
    *,
    sheet_name: str,
    reported_n_cols: int,
    rows: list[RawRow],
    column_hidden: list[bool],
    column_width: list[float | None],
) -> RawSheet:
    """Обрезает пустой хвост колонок справа и пустые строки снизу.

    Колонки/строки с реальными данными не трогает. После обрезки fingerprint
    по-прежнему требует ровно 17 колонок и строку «Итого» последней.
    """
    effective_cols = _effective_column_count(rows, reported_n_cols)
    if effective_cols < reported_n_cols:
        rows = [
            RawRow(index=row.index, outline_level=row.outline_level, values=row.values[:effective_cols])
            for row in rows
        ]
        column_hidden = column_hidden[:effective_cols]
        column_width = column_width[:effective_cols]
    else:
        effective_cols = reported_n_cols

    trimmed_rows = _trim_trailing_empty_rows(rows)
    return RawSheet(
        sheet_name=sheet_name,
        n_rows=len(trimmed_rows),
        n_cols=effective_cols,
        rows=tuple(trimmed_rows),
        column_hidden=tuple(column_hidden),
        column_width=tuple(column_width),
    )


def read_workbook(path: Path, sheet_name: str) -> RawSheet:
    """Читает лист `sheet_name` из `.xls` или `.xlsx` файла по расширению пути."""
    suffix = path.suffix.lower()
    if suffix == ".xls":
        return _read_xls(path, sheet_name)
    if suffix == ".xlsx":
        return _read_xlsx(path, sheet_name)
    raise ExcelReadError(f"Неподдерживаемое расширение файла: {suffix!r} (ожидается .xls или .xlsx)")


def _read_xls(path: Path, sheet_name: str) -> RawSheet:
    import xlrd

    try:
        book = xlrd.open_workbook(str(path), formatting_info=True)
    except Exception as exc:  # noqa: BLE001 — любая ошибка чтения => отклонение файла
        raise ExcelReadError(f"Не удалось прочитать .xls книгу: {exc}") from exc

    try:
        sheet = book.sheet_by_name(sheet_name)
    except Exception as exc:  # noqa: BLE001 — xlrd бросает разные типы при отсутствии листа
        raise SheetNotFoundError(f"Лист {sheet_name!r} не найден в книге") from exc

    n_cols = sheet.ncols
    rows: list[RawRow] = []
    for r in range(sheet.nrows):
        rowinfo = sheet.rowinfo_map.get(r)
        outline_level = rowinfo.outline_level if rowinfo is not None else 0
        values = tuple(sheet.cell_value(r, c) for c in range(n_cols))
        rows.append(RawRow(index=r, outline_level=outline_level, values=values))

    column_hidden = []
    column_width = []
    for c in range(n_cols):
        colinfo = sheet.colinfo_map.get(c)
        column_hidden.append(bool(colinfo.hidden) if colinfo is not None else False)
        column_width.append(float(colinfo.width) if colinfo is not None else None)

    return _trim_trailing_empty_columns(
        sheet_name=sheet.name,
        reported_n_cols=n_cols,
        rows=rows,
        column_hidden=column_hidden,
        column_width=column_width,
    )


def _read_xlsx(path: Path, sheet_name: str) -> RawSheet:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.exceptions import InvalidFileException

    try:
        workbook = openpyxl.load_workbook(str(path), data_only=True, read_only=False)
    except (InvalidFileException, OSError, KeyError, zipfile.BadZipFile) as exc:
        raise ExcelReadError(f"Не удалось прочитать .xlsx книгу: {exc}") from exc

    if sheet_name not in workbook.sheetnames:
        raise SheetNotFoundError(f"Лист {sheet_name!r} не найден в книге")
    worksheet = workbook[sheet_name]

    n_rows = worksheet.max_row or 0
    n_cols = worksheet.max_column or 0

    rows: list[RawRow] = []
    for excel_row in range(1, n_rows + 1):
        row_dim = worksheet.row_dimensions.get(excel_row)
        outline_level = row_dim.outlineLevel if row_dim is not None else 0
        values = tuple(
            worksheet.cell(row=excel_row, column=c).value for c in range(1, n_cols + 1)
        )
        rows.append(RawRow(index=excel_row - 1, outline_level=outline_level or 0, values=values))

    column_hidden = []
    column_width = []
    for c in range(1, n_cols + 1):
        letter = get_column_letter(c)
        col_dim = worksheet.column_dimensions.get(letter)
        column_hidden.append(bool(col_dim.hidden) if col_dim is not None else False)
        column_width.append(float(col_dim.width) if col_dim is not None and col_dim.width else None)

    return _trim_trailing_empty_columns(
        sheet_name=worksheet.title,
        reported_n_cols=n_cols,
        rows=rows,
        column_hidden=column_hidden,
        column_width=column_width,
    )
