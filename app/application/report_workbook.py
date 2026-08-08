"""CORE Excel workbook builder (Stage 4.2 corrective).

Deterministic openpyxl export: full additive metric set, formula-injection safe
labels, aggregates from raw L1 positions (including collisions), fixed ZIP/core
metadata for stable excel_sha256.
"""
from __future__ import annotations

import hashlib
import re
import zipfile
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal
from io import BytesIO
from typing import Callable, Sequence

from openpyxl import Workbook
from openpyxl.packaging.core import DocumentProperties
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook as WorkbookType
from openpyxl.worksheet.worksheet import Worksheet

from app.application.comparison_service import (
    ADDITIVE_METRICS,
    CycleComparison,
    MatchedEntity,
    MetricDelta,
    PositionSnapshot,
    abs_delta,
    percent_delta,
)

SHEET_SUMMARY = "Сводка"
SHEET_DEPARTMENTS = "Отделы"
SHEET_MANAGER_GROUPS = "ManagerGroup"
SHEET_COUNTERPARTIES = "Контрагенты"
SHEET_CONTRACTS = "Договоры"
SHEET_DOCUMENTS = "Документы"
SHEET_CHANGES = "Изменения"
SHEET_CONTROL = "Контроль"

SHEET_ORDER: tuple[str, ...] = (
    SHEET_SUMMARY,
    SHEET_DEPARTMENTS,
    SHEET_MANAGER_GROUPS,
    SHEET_COUNTERPARTIES,
    SHEET_CONTRACTS,
    SHEET_DOCUMENTS,
    SHEET_CHANGES,
    SHEET_CONTROL,
)

MONEY_FORMAT = "#,##0.00"
PCT_FORMAT = "0.00"
ND = "н/д"
CREATOR = "debitor-bot"
FIXED_TIMESTAMP = datetime(2020, 1, 1, 0, 0, 0)
ZIP_DATE_TIME = (2020, 1, 1, 0, 0, 0)
FIXED_CORE_ISO = "2020-01-01T00:00:00Z"

_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")

DEPARTMENT_LABELS: dict[str, str] = {
    "szfo_1": "СЗФО-1",
    "szfo_2": "СЗФО-2",
    "regional": "Региональный",
    "moscow": "Москва",
    "fokin": "Фокин",
}

CHANGE_NEW = "новая позиция"
CHANGE_CLOSED = "закрыта"
CHANGE_GROWTH = "рост долга"
CHANGE_DECLINE = "чистое снижение долга"
CHANGE_UNCHANGED = "без изменения total_debt"
CHANGE_AMBIGUOUS = "ambiguous collision"
CHANGE_OVERDUE = "изменение профиля просрочки"

YES = "да"
NO = "нет"


def metric_headers(metric: str) -> tuple[str, str, str, str]:
    return (
        f"{metric} current",
        f"{metric} previous",
        f"{metric} abs_delta",
        f"{metric} percent_delta",
    )


ALL_METRIC_HEADERS: tuple[str, ...] = tuple(
    header for metric in ADDITIVE_METRICS for header in metric_headers(metric)
)

SUMMARY_HEADERS: tuple[str, ...] = (
    "metric",
    "current",
    "previous",
    "abs_delta",
    "percent_delta",
)

DEPARTMENT_HEADERS: tuple[str, ...] = (
    "department_key",
    "department_label",
    "new",
    "closed",
    *ALL_METRIC_HEADERS,
)

MANAGER_GROUP_HEADERS: tuple[str, ...] = (
    "manager_group_id",
    "ManagerGroup",
    "department_key",
    "department_label",
    "new",
    "closed",
    *ALL_METRIC_HEADERS,
)

COUNTERPARTY_HEADERS: tuple[str, ...] = (
    "match_key",
    "department_key",
    "department_label",
    "manager_group_id",
    "ManagerGroup",
    "counterparty_id",
    "Контрагент",
    "Ярлык",
    "new",
    "closed",
    "класс изменения",
    "изменение профиля просрочки",
    *ALL_METRIC_HEADERS,
)

CONTRACT_HEADERS: tuple[str, ...] = (
    "match_key",
    "outline_level",
    "department_key",
    "department_label",
    "manager_group_id",
    "ManagerGroup",
    "counterparty_id",
    "Контрагент",
    "Договор",
    "Объект",
    "Ярлык",
    "new",
    "closed",
    "класс изменения",
    "изменение профиля просрочки",
    *ALL_METRIC_HEADERS,
)

DOCUMENT_HEADERS: tuple[str, ...] = (
    "match_key",
    "department_key",
    "department_label",
    "manager_group_id",
    "ManagerGroup",
    "counterparty_id",
    "Контрагент",
    "Договор",
    "Объект",
    "Ярлык",
    "new",
    "closed",
    "класс изменения",
    "изменение профиля просрочки",
    "переход корзины документа",
    *ALL_METRIC_HEADERS,
)

CHANGES_HEADERS: tuple[str, ...] = (
    "match_key",
    "outline_level",
    "new",
    "closed",
    "ambiguous",
    "collision_current_count",
    "collision_previous_count",
    "raw_labels",
    "department_key",
    "department_label",
    "manager_group_id",
    "ManagerGroup",
    "counterparty_id",
    "Контрагент",
    "Ярлык",
    "класс изменения",
    "изменение профиля просрочки",
    *ALL_METRIC_HEADERS,
)

CONTROL_HEADERS: tuple[str, ...] = (
    "тип",
    "имя",
    "ok",
    "diagnostic",
    "left",
    "right",
    "цикл",
    "match_key",
    "count",
    "raw_labels",
)

SHEET_HEADERS: dict[str, tuple[str, ...]] = {
    SHEET_SUMMARY: SUMMARY_HEADERS,
    SHEET_DEPARTMENTS: DEPARTMENT_HEADERS,
    SHEET_MANAGER_GROUPS: MANAGER_GROUP_HEADERS,
    SHEET_COUNTERPARTIES: COUNTERPARTY_HEADERS,
    SHEET_CONTRACTS: CONTRACT_HEADERS,
    SHEET_DOCUMENTS: DOCUMENT_HEADERS,
    SHEET_CHANGES: CHANGES_HEADERS,
    SHEET_CONTROL: CONTROL_HEADERS,
}


def sanitize_excel_text(value: str | None) -> str:
    """Neutralize formula injection for labels / names written as cell text."""
    if value is None:
        return ""
    text = str(value)
    if text.startswith(_FORMULA_PREFIXES):
        return "'" + text
    return text


def _abs_rank(delta: Decimal | None) -> Decimal:
    if delta is None:
        return Decimal("-1")
    return abs(delta)


def presence_flags(
    *,
    curr_present: bool,
    prev_present: bool,
) -> tuple[bool, bool]:
    return (curr_present and not prev_present, prev_present and not curr_present)


def change_class(entity: MatchedEntity) -> str:
    if entity.ambiguous:
        return CHANGE_AMBIGUOUS
    is_new, is_closed = presence_flags(
        curr_present=entity.current is not None,
        prev_present=entity.previous is not None,
    )
    if is_new:
        return CHANGE_NEW
    if is_closed:
        return CHANGE_CLOSED
    debt = entity.deltas["total_debt"].abs_delta
    if debt is None:
        base = CHANGE_UNCHANGED
    elif debt > 0:
        base = CHANGE_GROWTH
    elif debt < 0:
        base = CHANGE_DECLINE
    else:
        base = CHANGE_UNCHANGED
    if entity.overdue_profile_changed:
        return f"{base}; {CHANGE_OVERDUE}"
    return base


def _yes_no(flag: bool) -> str:
    return YES if flag else NO


def _sum_l1_metric(
    positions: Sequence[PositionSnapshot],
    metric: str,
    *,
    predicate: Callable[[PositionSnapshot], bool] | None = None,
) -> Decimal | None:
    total = Decimal("0")
    any_value = False
    for pos in positions:
        if pos.outline_level != 1:
            continue
        if predicate is not None and not predicate(pos):
            continue
        value = pos.metrics.get(metric)
        if value is None:
            continue
        total += value
        any_value = True
    return total if any_value else None


def _l1_present(
    positions: Sequence[PositionSnapshot],
    *,
    predicate: Callable[[PositionSnapshot], bool] | None = None,
) -> bool:
    for pos in positions:
        if pos.outline_level != 1:
            continue
        if predicate is not None and not predicate(pos):
            continue
        return True
    return False


def aggregate_metric_deltas(
    *,
    current_positions: Sequence[PositionSnapshot],
    previous_positions: Sequence[PositionSnapshot],
    predicate: Callable[[PositionSnapshot], bool] | None = None,
) -> dict[str, MetricDelta]:
    curr_present = _l1_present(current_positions, predicate=predicate)
    prev_present = _l1_present(previous_positions, predicate=predicate)
    out: dict[str, MetricDelta] = {}
    for metric in ADDITIVE_METRICS:
        curr_m = _sum_l1_metric(current_positions, metric, predicate=predicate)
        prev_m = _sum_l1_metric(previous_positions, metric, predicate=predicate)
        ad = abs_delta(
            curr_present=curr_present,
            prev_present=prev_present,
            curr_m=curr_m,
            prev_m=prev_m,
        )
        out[metric] = MetricDelta(
            current=curr_m if curr_present else None,
            previous=prev_m if prev_present else None,
            abs_delta=ad,
            percent_delta=percent_delta(ad, prev_m if prev_present else None),
        )
    return out


@dataclass(frozen=True)
class AggregateRow:
    key: str
    label: str
    department_key: str
    manager_group_id: int | None
    is_new: bool
    is_closed: bool
    deltas: dict[str, MetricDelta]


def build_department_rows(comparison: CycleComparison) -> list[AggregateRow]:
    depts = sorted(
        {
            p.department
            for p in (*comparison.current_positions, *comparison.previous_positions)
            if p.outline_level == 1
        }
    )
    rows: list[AggregateRow] = []
    for dept in depts:
        pred = lambda s, d=dept: s.department == d
        deltas = aggregate_metric_deltas(
            current_positions=comparison.current_positions,
            previous_positions=comparison.previous_positions,
            predicate=pred,
        )
        curr_present = _l1_present(comparison.current_positions, predicate=pred)
        prev_present = _l1_present(comparison.previous_positions, predicate=pred)
        is_new, is_closed = presence_flags(
            curr_present=curr_present, prev_present=prev_present
        )
        rows.append(
            AggregateRow(
                key=dept,
                label=DEPARTMENT_LABELS.get(dept, dept),
                department_key=dept,
                manager_group_id=None,
                is_new=is_new,
                is_closed=is_closed,
                deltas=deltas,
            )
        )
    return sorted(
        rows,
        key=lambda r: (-_abs_rank(r.deltas["total_debt"].abs_delta), r.key),
    )


def build_manager_group_rows(comparison: CycleComparison) -> list[AggregateRow]:
    groups: dict[int, tuple[str, str]] = {}
    for pos in (*comparison.current_positions, *comparison.previous_positions):
        if pos.outline_level != 1:
            continue
        groups[pos.manager_group_id] = (
            pos.manager_group_name or str(pos.manager_group_id),
            pos.department,
        )
    rows: list[AggregateRow] = []
    for mg_id in sorted(groups):
        pred = lambda s, mid=mg_id: s.manager_group_id == mid
        deltas = aggregate_metric_deltas(
            current_positions=comparison.current_positions,
            previous_positions=comparison.previous_positions,
            predicate=pred,
        )
        curr_present = _l1_present(comparison.current_positions, predicate=pred)
        prev_present = _l1_present(comparison.previous_positions, predicate=pred)
        is_new, is_closed = presence_flags(
            curr_present=curr_present, prev_present=prev_present
        )
        label, dept = groups[mg_id]
        rows.append(
            AggregateRow(
                key=str(mg_id),
                label=label,
                department_key=dept,
                manager_group_id=mg_id,
                is_new=is_new,
                is_closed=is_closed,
                deltas=deltas,
            )
        )
    return sorted(
        rows,
        key=lambda r: (-_abs_rank(r.deltas["total_debt"].abs_delta), r.key),
    )


def _sort_entities(entities: Sequence[MatchedEntity]) -> list[MatchedEntity]:
    return sorted(
        entities,
        key=lambda e: (
            -_abs_rank(e.deltas["total_debt"].abs_delta),
            e.match_key,
        ),
    )


def _positions_by_id(comparison: CycleComparison) -> dict[int, PositionSnapshot]:
    out: dict[int, PositionSnapshot] = {}
    for snap in (*comparison.current_positions, *comparison.previous_positions):
        out[snap.id] = snap
    return out


def _ancestor_label(
    snap: PositionSnapshot | None,
    *,
    by_id: dict[int, PositionSnapshot],
    outline_level: int,
) -> str:
    cur = snap
    while cur is not None:
        if cur.outline_level == outline_level:
            return cur.raw_label
        if cur.parent_position_id is None:
            break
        cur = by_id.get(cur.parent_position_id)
    return ""


def _collision_raw_labels(comparison: CycleComparison, match_key: str) -> str:
    labels: list[str] = []
    for collision in comparison.collisions:
        if collision.match_key != match_key:
            continue
        labels.extend(collision.raw_labels)
    # Preserve order, drop duplicates.
    seen: set[str] = set()
    unique: list[str] = []
    for label in labels:
        if label in seen:
            continue
        seen.add(label)
        unique.append(label)
    return " | ".join(unique)


def _display_from_positions(
    comparison: CycleComparison, match_key: str
) -> PositionSnapshot | None:
    for pos in comparison.current_positions:
        if pos.match_key == match_key:
            return pos
    for pos in comparison.previous_positions:
        if pos.match_key == match_key:
            return pos
    return None


@dataclass(frozen=True)
class EntityDisplay:
    department_key: str
    department_label: str
    manager_group_id: int | None
    manager_group_name: str
    counterparty_id: int | None
    counterparty_name: str
    raw_label: str


def _entity_display(
    entity: MatchedEntity, comparison: CycleComparison
) -> EntityDisplay:
    snap = entity.current or entity.previous
    if snap is None:
        snap = _display_from_positions(comparison, entity.match_key)
    if snap is None:
        return EntityDisplay(
            department_key="",
            department_label="",
            manager_group_id=None,
            manager_group_name="",
            counterparty_id=None,
            counterparty_name="",
            raw_label=_collision_raw_labels(comparison, entity.match_key),
        )
    return EntityDisplay(
        department_key=snap.department,
        department_label=DEPARTMENT_LABELS.get(snap.department, snap.department),
        manager_group_id=snap.manager_group_id,
        manager_group_name=snap.manager_group_name or str(snap.manager_group_id),
        counterparty_id=snap.counterparty_id,
        counterparty_name=snap.counterparty_name or str(snap.counterparty_id),
        raw_label=snap.raw_label,
    )


def _write_header(ws: Worksheet, headers: Sequence[str]) -> None:
    for col, title in enumerate(headers, start=1):
        ws.cell(1, col, title)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _set_money(cell) -> None:
    cell.number_format = MONEY_FORMAT


def _write_money(ws: Worksheet, row: int, col: int, value: Decimal | None):
    cell = ws.cell(row, col, value)
    cell.number_format = MONEY_FORMAT
    return cell


def _write_pct(ws: Worksheet, row: int, col: int, value: Decimal | None) -> None:
    if value is None:
        ws.cell(row, col, ND)
        return
    cell = ws.cell(row, col, value)
    cell.number_format = PCT_FORMAT


def _write_metric_block(
    ws: Worksheet, row: int, start_col: int, deltas: dict[str, MetricDelta]
) -> int:
    col = start_col
    for metric in ADDITIVE_METRICS:
        delta = deltas[metric]
        _write_money(ws, row, col, delta.current)
        col += 1
        _write_money(ws, row, col, delta.previous)
        col += 1
        _write_money(ws, row, col, delta.abs_delta)
        col += 1
        _write_pct(ws, row, col, delta.percent_delta)
        col += 1
    return col


def _apply_column_widths(ws: Worksheet, widths: Sequence[float]) -> None:
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _finalize_table(ws: Worksheet, headers: Sequence[str], row_count: int) -> None:
    last_row = max(row_count + 1, 1)
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"
    _apply_column_widths(ws, [16] * len(headers))


def _normalize_xlsx_bytes(raw: bytes) -> bytes:
    """Rewrite ZIP entry timestamps and core created/modified for stable SHA."""
    src = zipfile.ZipFile(BytesIO(raw), "r")
    out = BytesIO()
    with zipfile.ZipFile(out, "w", compression=zipfile.ZIP_DEFLATED) as dst:
        for info in sorted(src.infolist(), key=lambda item: item.filename):
            data = src.read(info.filename)
            if info.filename == "docProps/core.xml":
                text = data.decode("utf-8")
                text = re.sub(
                    r"(<(?:dcterms:)?created[^>]*>)([^<]+)(</(?:dcterms:)?created>)",
                    rf"\g<1>{FIXED_CORE_ISO}\g<3>",
                    text,
                )
                text = re.sub(
                    r"(<(?:dcterms:)?modified[^>]*>)([^<]+)(</(?:dcterms:)?modified>)",
                    rf"\g<1>{FIXED_CORE_ISO}\g<3>",
                    text,
                )
                data = text.encode("utf-8")
            new_info = zipfile.ZipInfo(filename=info.filename, date_time=ZIP_DATE_TIME)
            new_info.compress_type = zipfile.ZIP_DEFLATED
            dst.writestr(new_info, data)
    src.close()
    return out.getvalue()


def _prepare_workbook() -> WorkbookType:
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    for name in SHEET_ORDER:
        wb.create_sheet(name)
    wb.properties = DocumentProperties(
        creator=CREATOR,
        lastModifiedBy=CREATOR,
        created=FIXED_TIMESTAMP,
        modified=FIXED_TIMESTAMP,
    )
    return wb


def _fill_summary(ws: Worksheet, comparison: CycleComparison) -> None:
    _write_header(ws, SUMMARY_HEADERS)
    company = aggregate_metric_deltas(
        current_positions=comparison.current_positions,
        previous_positions=comparison.previous_positions,
    )
    for r_idx, metric in enumerate(ADDITIVE_METRICS, start=2):
        delta = company[metric]
        ws.cell(r_idx, 1, metric)
        _write_money(ws, r_idx, 2, delta.current)
        _write_money(ws, r_idx, 3, delta.previous)
        _write_money(ws, r_idx, 4, delta.abs_delta)
        _write_pct(ws, r_idx, 5, delta.percent_delta)

    # Cycle metadata beside the metric table (not part of AutoFilter range).
    meta = [
        ("current_cycle_id", comparison.current_cycle_id),
        ("current_report_date", comparison.current_report_date.isoformat()),
        (
            "previous_cycle_id",
            comparison.previous_cycle_id
            if comparison.previous_cycle_id is not None
            else ND,
        ),
        (
            "previous_report_date",
            comparison.previous_report_date.isoformat()
            if comparison.previous_report_date
            else ND,
        ),
        ("entity_count", len(comparison.entities)),
        ("ambiguous_key_count", len(comparison.ambiguous_keys)),
        ("collision_count", len(comparison.collisions)),
        (
            "overdue_profile_change_count",
            sum(1 for e in comparison.entities if e.overdue_profile_changed),
        ),
        (
            "control_failure_count",
            sum(
                1
                for c in comparison.control_equalities
                if not c.ok and not c.diagnostic
            ),
        ),
    ]
    ws.cell(1, 7, "meta")
    ws.cell(1, 8, "value")
    for idx, (name, value) in enumerate(meta, start=2):
        ws.cell(idx, 7, name)
        ws.cell(idx, 8, value)
    _finalize_table(ws, SUMMARY_HEADERS, len(ADDITIVE_METRICS))


def _fill_department_sheet(ws: Worksheet, comparison: CycleComparison) -> None:
    _write_header(ws, DEPARTMENT_HEADERS)
    rows = build_department_rows(comparison)
    for r_idx, row in enumerate(rows, start=2):
        ws.cell(r_idx, 1, row.department_key)
        ws.cell(r_idx, 2, sanitize_excel_text(row.label))
        ws.cell(r_idx, 3, _yes_no(row.is_new))
        ws.cell(r_idx, 4, _yes_no(row.is_closed))
        _write_metric_block(ws, r_idx, 5, row.deltas)
    _finalize_table(ws, DEPARTMENT_HEADERS, len(rows))


def _fill_manager_group_sheet(ws: Worksheet, comparison: CycleComparison) -> None:
    _write_header(ws, MANAGER_GROUP_HEADERS)
    rows = build_manager_group_rows(comparison)
    for r_idx, row in enumerate(rows, start=2):
        ws.cell(r_idx, 1, row.manager_group_id)
        ws.cell(r_idx, 2, sanitize_excel_text(row.label))
        ws.cell(r_idx, 3, row.department_key)
        ws.cell(
            r_idx,
            4,
            sanitize_excel_text(
                DEPARTMENT_LABELS.get(row.department_key, row.department_key)
            ),
        )
        ws.cell(r_idx, 5, _yes_no(row.is_new))
        ws.cell(r_idx, 6, _yes_no(row.is_closed))
        _write_metric_block(ws, r_idx, 7, row.deltas)
    _finalize_table(ws, MANAGER_GROUP_HEADERS, len(rows))


def _fill_counterparty_sheet(
    ws: Worksheet,
    entities: Sequence[MatchedEntity],
    comparison: CycleComparison,
) -> None:
    _write_header(ws, COUNTERPARTY_HEADERS)
    sorted_entities = _sort_entities(entities)
    for r_idx, entity in enumerate(sorted_entities, start=2):
        disp = _entity_display(entity, comparison)
        is_new, is_closed = presence_flags(
            curr_present=entity.current is not None,
            prev_present=entity.previous is not None,
        )
        if entity.ambiguous:
            is_new, is_closed = False, False
        ws.cell(r_idx, 1, entity.match_key)
        ws.cell(r_idx, 2, disp.department_key)
        ws.cell(r_idx, 3, sanitize_excel_text(disp.department_label))
        ws.cell(r_idx, 4, disp.manager_group_id)
        ws.cell(r_idx, 5, sanitize_excel_text(disp.manager_group_name))
        ws.cell(r_idx, 6, disp.counterparty_id)
        ws.cell(r_idx, 7, sanitize_excel_text(disp.counterparty_name))
        ws.cell(r_idx, 8, sanitize_excel_text(disp.raw_label))
        ws.cell(r_idx, 9, _yes_no(is_new))
        ws.cell(r_idx, 10, _yes_no(is_closed))
        ws.cell(r_idx, 11, change_class(entity))
        ws.cell(r_idx, 12, _yes_no(entity.overdue_profile_changed))
        _write_metric_block(ws, r_idx, 13, entity.deltas)
    _finalize_table(ws, COUNTERPARTY_HEADERS, len(sorted_entities))


def _fill_contract_sheet(
    ws: Worksheet,
    entities: Sequence[MatchedEntity],
    comparison: CycleComparison,
    by_id: dict[int, PositionSnapshot],
) -> None:
    _write_header(ws, CONTRACT_HEADERS)
    sorted_entities = _sort_entities(entities)
    for r_idx, entity in enumerate(sorted_entities, start=2):
        disp = _entity_display(entity, comparison)
        snap = entity.current or entity.previous or _display_from_positions(
            comparison, entity.match_key
        )
        is_new, is_closed = presence_flags(
            curr_present=entity.current is not None,
            prev_present=entity.previous is not None,
        )
        if entity.ambiguous:
            is_new, is_closed = False, False
        contract = _ancestor_label(snap, by_id=by_id, outline_level=2)
        obj = _ancestor_label(snap, by_id=by_id, outline_level=3)
        if snap is not None and snap.outline_level == 2:
            contract = snap.raw_label
        if snap is not None and snap.outline_level == 3:
            obj = snap.raw_label
        ws.cell(r_idx, 1, entity.match_key)
        ws.cell(r_idx, 2, entity.outline_level)
        ws.cell(r_idx, 3, disp.department_key)
        ws.cell(r_idx, 4, sanitize_excel_text(disp.department_label))
        ws.cell(r_idx, 5, disp.manager_group_id)
        ws.cell(r_idx, 6, sanitize_excel_text(disp.manager_group_name))
        ws.cell(r_idx, 7, disp.counterparty_id)
        ws.cell(r_idx, 8, sanitize_excel_text(disp.counterparty_name))
        ws.cell(r_idx, 9, sanitize_excel_text(contract))
        ws.cell(r_idx, 10, sanitize_excel_text(obj))
        ws.cell(r_idx, 11, sanitize_excel_text(disp.raw_label))
        ws.cell(r_idx, 12, _yes_no(is_new))
        ws.cell(r_idx, 13, _yes_no(is_closed))
        ws.cell(r_idx, 14, change_class(entity))
        ws.cell(r_idx, 15, _yes_no(entity.overdue_profile_changed))
        _write_metric_block(ws, r_idx, 16, entity.deltas)
    _finalize_table(ws, CONTRACT_HEADERS, len(sorted_entities))


def _fill_document_sheet(
    ws: Worksheet,
    entities: Sequence[MatchedEntity],
    comparison: CycleComparison,
    by_id: dict[int, PositionSnapshot],
) -> None:
    _write_header(ws, DOCUMENT_HEADERS)
    sorted_entities = _sort_entities(entities)
    for r_idx, entity in enumerate(sorted_entities, start=2):
        disp = _entity_display(entity, comparison)
        snap = entity.current or entity.previous or _display_from_positions(
            comparison, entity.match_key
        )
        is_new, is_closed = presence_flags(
            curr_present=entity.current is not None,
            prev_present=entity.previous is not None,
        )
        if entity.ambiguous:
            is_new, is_closed = False, False
        contract = _ancestor_label(snap, by_id=by_id, outline_level=2)
        obj = _ancestor_label(snap, by_id=by_id, outline_level=3)
        ws.cell(r_idx, 1, entity.match_key)
        ws.cell(r_idx, 2, disp.department_key)
        ws.cell(r_idx, 3, sanitize_excel_text(disp.department_label))
        ws.cell(r_idx, 4, disp.manager_group_id)
        ws.cell(r_idx, 5, sanitize_excel_text(disp.manager_group_name))
        ws.cell(r_idx, 6, disp.counterparty_id)
        ws.cell(r_idx, 7, sanitize_excel_text(disp.counterparty_name))
        ws.cell(r_idx, 8, sanitize_excel_text(contract))
        ws.cell(r_idx, 9, sanitize_excel_text(obj))
        ws.cell(r_idx, 10, sanitize_excel_text(disp.raw_label))
        ws.cell(r_idx, 11, _yes_no(is_new))
        ws.cell(r_idx, 12, _yes_no(is_closed))
        ws.cell(r_idx, 13, change_class(entity))
        ws.cell(r_idx, 14, _yes_no(entity.overdue_profile_changed))
        if entity.document_bucket_transition is None:
            ws.cell(r_idx, 15, "")
        else:
            frm, to = entity.document_bucket_transition
            ws.cell(r_idx, 15, f"{frm} → {to}")
        _write_metric_block(ws, r_idx, 16, entity.deltas)
    _finalize_table(ws, DOCUMENT_HEADERS, len(sorted_entities))


def _fill_changes(ws: Worksheet, comparison: CycleComparison) -> None:
    _write_header(ws, CHANGES_HEADERS)
    interesting: list[MatchedEntity] = []
    for entity in comparison.entities:
        if entity.ambiguous:
            interesting.append(entity)
            continue
        debt = entity.deltas["total_debt"].abs_delta
        if entity.current is None or entity.previous is None:
            interesting.append(entity)
            continue
        if debt is not None and debt != 0:
            interesting.append(entity)
            continue
        if entity.overdue_profile_changed:
            interesting.append(entity)
    interesting = _sort_entities(interesting)
    for r_idx, entity in enumerate(interesting, start=2):
        disp = _entity_display(entity, comparison)
        is_new, is_closed = presence_flags(
            curr_present=entity.current is not None,
            prev_present=entity.previous is not None,
        )
        if entity.ambiguous:
            is_new, is_closed = False, False
        ws.cell(r_idx, 1, entity.match_key)
        ws.cell(r_idx, 2, entity.outline_level)
        ws.cell(r_idx, 3, _yes_no(is_new))
        ws.cell(r_idx, 4, _yes_no(is_closed))
        ws.cell(r_idx, 5, _yes_no(entity.ambiguous))
        ws.cell(r_idx, 6, entity.collision_current_count)
        ws.cell(r_idx, 7, entity.collision_previous_count)
        ws.cell(
            r_idx,
            8,
            " | ".join(
                sanitize_excel_text(part)
                for part in _collision_raw_labels(comparison, entity.match_key).split(
                    " | "
                )
                if part
            ),
        )
        ws.cell(r_idx, 9, disp.department_key)
        ws.cell(r_idx, 10, sanitize_excel_text(disp.department_label))
        ws.cell(r_idx, 11, disp.manager_group_id)
        ws.cell(r_idx, 12, sanitize_excel_text(disp.manager_group_name))
        ws.cell(r_idx, 13, disp.counterparty_id)
        ws.cell(r_idx, 14, sanitize_excel_text(disp.counterparty_name))
        ws.cell(r_idx, 15, sanitize_excel_text(disp.raw_label))
        ws.cell(r_idx, 16, change_class(entity))
        ws.cell(r_idx, 17, _yes_no(entity.overdue_profile_changed))
        _write_metric_block(ws, r_idx, 18, entity.deltas)
    _finalize_table(ws, CHANGES_HEADERS, len(interesting))


def _fill_control(ws: Worksheet, comparison: CycleComparison) -> None:
    _write_header(ws, CONTROL_HEADERS)
    r_idx = 2
    for check in comparison.control_equalities:
        ws.cell(r_idx, 1, "equality")
        ws.cell(r_idx, 2, check.name)
        ws.cell(r_idx, 3, "ok" if check.ok else "FAIL")
        ws.cell(r_idx, 4, _yes_no(check.diagnostic))
        _write_money(ws, r_idx, 5, check.left)
        _write_money(ws, r_idx, 6, check.right)
        r_idx += 1
    for collision in comparison.collisions:
        ws.cell(r_idx, 1, "collision")
        ws.cell(r_idx, 2, "match_key_collision")
        ws.cell(r_idx, 3, "FAIL")
        ws.cell(r_idx, 4, NO)
        ws.cell(r_idx, 7, collision.cycle_side)
        ws.cell(r_idx, 8, collision.match_key)
        ws.cell(r_idx, 9, collision.count)
        ws.cell(
            r_idx,
            10,
            " | ".join(sanitize_excel_text(label) for label in collision.raw_labels),
        )
        r_idx += 1
    _finalize_table(ws, CONTROL_HEADERS, r_idx - 2)


def build_core_workbook(comparison: CycleComparison) -> WorkbookType:
    wb = _prepare_workbook()
    by_id = _positions_by_id(comparison)

    _fill_summary(wb[SHEET_SUMMARY], comparison)
    _fill_department_sheet(wb[SHEET_DEPARTMENTS], comparison)
    _fill_manager_group_sheet(wb[SHEET_MANAGER_GROUPS], comparison)

    counterparties = [e for e in comparison.entities if e.outline_level == 1]
    contracts_and_objects = [
        e for e in comparison.entities if e.outline_level in (2, 3)
    ]
    documents = [e for e in comparison.entities if e.outline_level == 4]

    _fill_counterparty_sheet(
        wb[SHEET_COUNTERPARTIES], counterparties, comparison
    )
    _fill_contract_sheet(
        wb[SHEET_CONTRACTS], contracts_and_objects, comparison, by_id
    )
    _fill_document_sheet(wb[SHEET_DOCUMENTS], documents, comparison, by_id)
    _fill_changes(wb[SHEET_CHANGES], comparison)
    _fill_control(wb[SHEET_CONTROL], comparison)
    return wb


def workbook_to_bytes(workbook: WorkbookType) -> bytes:
    bio = BytesIO()
    workbook.save(bio)
    return _normalize_xlsx_bytes(bio.getvalue())


def build_core_excel_bytes(comparison: CycleComparison) -> tuple[bytes, str]:
    """Return (xlsx_bytes, sha256_hex) for a CORE artifact."""
    wb = build_core_workbook(comparison)
    raw = workbook_to_bytes(wb)
    digest = hashlib.sha256(raw).hexdigest()
    return raw, digest
