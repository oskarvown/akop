"""Build ENRICHED workbook from immutable CORE bytes + Комментарии sheet."""
from __future__ import annotations

import hashlib
from io import BytesIO
from typing import Any, Mapping, Sequence

from openpyxl import load_workbook

from app.application.report_workbook import FIXED_CORE_ISO, FIXED_TIMESTAMP, ZIP_DATE_TIME
from openpyxl.packaging.core import DocumentProperties
import zipfile

SHEET_COMMENTS = "Комментарии"

COMMENT_HEADERS: tuple[str, ...] = (
    "debt_position_id",
    "source_file_id",
    "row_order",
    "department",
    "manager_group",
    "counterparty_label",
    "outline_level",
    "comment_raw",
    "source",
    "analysis_status",
    "confidence",
    "mentioned_date",
    "mentioned_amount",
    "action",
    "reason",
    "responsible_person",
    "summary",
)


def build_enriched_excel_bytes(
    core_excel_bytes: bytes,
    *,
    snapshot_rows: Sequence[Mapping[str, Any]],
    analyses_by_position: Mapping[int, Mapping[str, Any]],
) -> tuple[bytes, str]:
    """Clone CORE workbook and append Комментарии from frozen snapshot + analyses."""
    workbook = load_workbook(BytesIO(core_excel_bytes))
    if SHEET_COMMENTS in workbook.sheetnames:
        del workbook[SHEET_COMMENTS]
    sheet = workbook.create_sheet(SHEET_COMMENTS)
    for col, header in enumerate(COMMENT_HEADERS, start=1):
        sheet.cell(1, col, header)
    for row_idx, snap in enumerate(snapshot_rows, start=2):
        position_id = int(snap["debt_position_id"])
        analysis = analyses_by_position.get(position_id, {})
        values = [
            snap.get("debt_position_id"),
            snap.get("source_file_id"),
            snap.get("row_order"),
            snap.get("department"),
            snap.get("manager_group"),
            snap.get("counterparty_label"),
            snap.get("outline_level"),
            snap.get("comment_raw"),
            analysis.get("source"),
            analysis.get("analysis_status"),
            analysis.get("confidence"),
            analysis.get("mentioned_date"),
            analysis.get("mentioned_amount"),
            analysis.get("action"),
            analysis.get("reason"),
            analysis.get("responsible_person"),
            analysis.get("summary"),
        ]
        for col, value in enumerate(values, start=1):
            sheet.cell(row_idx, col, value)

    props = DocumentProperties()
    props.creator = "debitor-bot"
    props.lastModifiedBy = "debitor-bot"
    props.created = FIXED_TIMESTAMP
    props.modified = FIXED_TIMESTAMP
    workbook.properties = props

    buffer = BytesIO()
    workbook.save(buffer)
    raw = buffer.getvalue()
    stabilized = _stabilize_zip_timestamps(raw)
    digest = hashlib.sha256(stabilized).hexdigest()
    return stabilized, digest


def _stabilize_zip_timestamps(payload: bytes) -> bytes:
    """Match CORE deterministic ZIP date/time where possible."""
    src = zipfile.ZipFile(BytesIO(payload), "r")
    out_buf = BytesIO()
    with zipfile.ZipFile(out_buf, "w", compression=zipfile.ZIP_DEFLATED) as dst:
        for info in src.infolist():
            data = src.read(info.filename)
            if info.filename == "docProps/core.xml":
                # Keep fixed ISO timestamps if present; leave as-is otherwise.
                text = data.decode("utf-8")
                text = text.replace(FIXED_CORE_ISO, FIXED_CORE_ISO)
                data = text.encode("utf-8")
            new_info = zipfile.ZipInfo(info.filename, date_time=ZIP_DATE_TIME)
            new_info.compress_type = zipfile.ZIP_DEFLATED
            dst.writestr(new_info, data)
    src.close()
    return out_buf.getvalue()
