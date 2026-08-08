"""Stage 4.4 comment enrichment + ENRICHED delivery integration tests."""
from __future__ import annotations

import datetime as dt
import hashlib
import uuid
from dataclasses import replace
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import AsyncMock

import pytest
from sqlalchemy import func, select, text
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker

from app.application.audit_service import add_source_file_atomic
from app.application.comment_enrichment_service import (
    EnrichmentConflictError,
    claim_enrichment_job,
    complete_enrichment,
    enqueue_enrichment_job,
    fail_enrichment,
    record_comment_analysis,
    retry_terminal_enrichment,
    run_claimed_enrichment,
)
from app.application.report_service import (
    claim_report_build,
    complete_report_build,
    run_claimed_build,
)
from app.bot.scheduler.report_scheduler import ReportScheduler
from app.domain.enums import Department
from app.domain.models import (
    AuditArtifact,
    AuditArtifactKind,
    AuditCycleStatus,
    AuditReport,
    AuditReportStatus,
    CommentAnalysis,
    CommentAnalysisConfidence,
    CommentAnalysisSource,
    CommentAnalysisStatus,
    CommentEnrichmentJob,
    CommentEnrichmentJobStatus,
    DebtPosition,
    ReportDelivery,
    ReportDeliveryKind,
    ReportDeliveryStatus,
)
from app.infrastructure.excel.validator import (
    ValidationResult,
    validate_confirmed_template_file,
)
from app.infrastructure.llm.openrouter_client import (
    LlmCommentFacts,
    OpenRouterTransientError,
)

FIXTURE = (
    Path(__file__).resolve().parents[1]
    / "fixtures"
    / "regional"
    / "regional_valid_basic.xls"
)
NOTIFY_CHAT_ID = 743971617


@pytest.fixture
def valid_result() -> ValidationResult:
    result = validate_confirmed_template_file(FIXTURE)
    assert result.is_valid and result.parsed is not None
    return result


def result_for_date(result: ValidationResult, report_date: dt.date) -> ValidationResult:
    assert result.parsed is not None
    return replace(result, parsed=replace(result.parsed, report_date=report_date))


def _settings(**overrides: object) -> SimpleNamespace:
    base = {
        "report_build_claim_ttl_seconds": 300,
        "report_build_max_attempts": 5,
        "report_build_backoff_seconds": 1,
        "report_scheduler_poll_seconds": 30,
        "report_delivery_claim_ttl_seconds": 300,
        "report_delivery_send_timeout_seconds": 30,
        "report_delivery_max_attempts": 5,
        "report_delivery_backoff_seconds": 1,
        "report_delivery_max_file_bytes": 52428800,
        "report_delivery_batch_size": 10,
        "openrouter_api_key": None,
        "openrouter_model": None,
        "openrouter_base_url": "https://openrouter.ai/api/v1",
        "openrouter_timeout_seconds": 5,
        "openrouter_max_retries": 1,
        "comment_enrichment_claim_ttl_seconds": 60,
        "comment_enrichment_run_timeout_seconds": 30,
        "comment_enrichment_max_attempts": 2,
        "comment_enrichment_backoff_seconds": 1,
        "comment_enrichment_batch_size": 10,
        "comment_parser_version": "1",
        "comment_prompt_version": "1",
        "comment_schema_version_llm": "1",
        "comment_redaction_version": "1",
        "llm_api_key": None,
        "llm_model": None,
    }
    base.update(overrides)
    return SimpleNamespace(**base)


async def complete_cycle(
    session: AsyncSession,
    result: ValidationResult,
    *,
    report_date: dt.date,
    sha_prefix: str,
) -> int:
    dated = result_for_date(result, report_date)
    last = None
    for index, department in enumerate(Department):
        last = await add_source_file_atomic(
            session,
            result=dated,
            department=department,
            sha256=f"{sha_prefix}-{index}",
            original_filename=f"{sha_prefix}-{index}.xls",
            report_date=report_date,
            notification_chat_id=NOTIFY_CHAT_ID,
        )
    assert last is not None
    assert last.status == AuditCycleStatus.COMPLETED
    return last.cycle_id


async def build_ready_core(
    maker: async_sessionmaker[AsyncSession], *, cycle_id: int
) -> tuple[int, int, object]:
    settings = _settings()
    async with maker() as session:
        report = await session.scalar(
            select(AuditReport).where(AuditReport.audit_cycle_id == cycle_id)
        )
        assert report is not None
        report_id = report.id
    async with maker() as session:
        claim = await claim_report_build(
            session, report_id=report_id, settings=settings
        )
    assert claim is not None
    async with maker() as session:
        result = await run_claimed_build(session, claim=claim)
    async with maker() as session:
        await complete_report_build(
            session,
            report_id=report_id,
            claim_token=claim.claim_token,
            summary_json=result.summary_json,
            excel_bytes=result.excel_bytes,
            excel_sha256=result.excel_sha256,
        )
    async with maker() as session:
        artifact = await session.scalar(
            select(AuditArtifact).where(
                AuditArtifact.audit_report_id == report_id,
                AuditArtifact.kind == AuditArtifactKind.CORE,
            )
        )
        assert artifact is not None
        return report_id, int(artifact.id), result


@pytest.mark.asyncio
async def test_enqueue_after_core_failure_keeps_core(
    stage3_session_maker: async_sessionmaker[AsyncSession],
    valid_result: ValidationResult,
) -> None:
    async with stage3_session_maker() as session:
        cycle_id = await complete_cycle(
            session,
            valid_result,
            report_date=dt.date(2026, 11, 1),
            sha_prefix="s44-iso",
        )
        await session.commit()
    report_id, artifact_id, _ = await build_ready_core(
        stage3_session_maker, cycle_id=cycle_id
    )
    settings = _settings()
    async with stage3_session_maker() as session:
        job_id = await enqueue_enrichment_job(
            session, report_id=report_id, settings=settings
        )
    assert job_id is not None
    async with stage3_session_maker() as session:
        report = await session.get(AuditReport, report_id)
        assert report is not None
        assert report.status == AuditReportStatus.READY
        artifact = await session.get(AuditArtifact, artifact_id)
        assert artifact is not None
        delivery = await session.scalar(
            select(ReportDelivery).where(
                ReportDelivery.audit_artifact_id == artifact_id,
                ReportDelivery.kind == ReportDeliveryKind.AUTOMATIC,
            )
        )
        assert delivery is not None


@pytest.mark.asyncio
async def test_zero_comments_skipped_no_enriched(
    stage3_session_maker: async_sessionmaker[AsyncSession],
    valid_result: ValidationResult,
) -> None:
    async with stage3_session_maker() as session:
        cycle_id = await complete_cycle(
            session,
            valid_result,
            report_date=dt.date(2026, 11, 2),
            sha_prefix="s44-zero",
        )
        await session.commit()
    report_id, _, _ = await build_ready_core(stage3_session_maker, cycle_id=cycle_id)
    settings = _settings()
    async with stage3_session_maker() as session:
        # Clear all comments so snapshot is empty
        await session.execute(
            text("UPDATE debt_positions SET comment_raw = NULL")
        )
        await session.commit()
    async with stage3_session_maker() as session:
        job_id = await enqueue_enrichment_job(
            session, report_id=report_id, settings=settings
        )
    assert job_id is not None
    async with stage3_session_maker() as session:
        job = await session.get(CommentEnrichmentJob, job_id)
        assert job is not None
        assert job.status == CommentEnrichmentJobStatus.SKIPPED
        enriched = await session.scalar(
            select(AuditArtifact.id).where(
                AuditArtifact.audit_report_id == report_id,
                AuditArtifact.kind == AuditArtifactKind.ENRICHED,
            )
        )
        assert enriched is None


@pytest.mark.asyncio
async def test_deterministic_enrichment_builds_enriched_and_auto(
    stage3_session_maker: async_sessionmaker[AsyncSession],
    valid_result: ValidationResult,
) -> None:
    async with stage3_session_maker() as session:
        cycle_id = await complete_cycle(
            session,
            valid_result,
            report_date=dt.date(2026, 11, 3),
            sha_prefix="s44-det",
        )
        await session.commit()
    report_id, core_id, _ = await build_ready_core(
        stage3_session_maker, cycle_id=cycle_id
    )
    settings = _settings()
    async with stage3_session_maker() as session:
        # Ensure at least one deterministic-parseable comment
        pos = await session.scalar(select(DebtPosition).limit(1))
        assert pos is not None
        pos.comment_raw = "Оплата 20.11 на 5000"
        await session.commit()

    async with stage3_session_maker() as session:
        job_id = await enqueue_enrichment_job(
            session, report_id=report_id, settings=settings
        )
    assert job_id is not None
    async with stage3_session_maker() as session:
        claim = await claim_enrichment_job(
            session, job_id=job_id, settings=settings
        )
    assert claim is not None
    result = await run_claimed_enrichment(
        stage3_session_maker,
        claim=claim,
        settings=settings,
        llm_client=None,
    )
    assert result is not None
    assert result.revision == 1
    async with stage3_session_maker() as session:
        job = await session.get(CommentEnrichmentJob, job_id)
        assert job is not None
        assert job.status == CommentEnrichmentJobStatus.READY
        enriched = await session.scalar(
            select(AuditArtifact).where(
                AuditArtifact.enrichment_job_id == job_id
            )
        )
        assert enriched is not None
        assert enriched.revision == 1
        delivery = await session.scalar(
            select(ReportDelivery).where(
                ReportDelivery.audit_artifact_id == enriched.id,
                ReportDelivery.kind == ReportDeliveryKind.AUTOMATIC,
            )
        )
        assert delivery is not None
        core = await session.get(AuditArtifact, core_id)
        assert core is not None
        # Financial sheets parity: same sheet names except Комментарии
        from openpyxl import load_workbook
        from io import BytesIO

        core_wb = load_workbook(BytesIO(core.excel_bytes))
        enr_wb = load_workbook(BytesIO(enriched.excel_bytes))
        assert "Комментарии" in enr_wb.sheetnames
        for name in core_wb.sheetnames:
            assert name in enr_wb.sheetnames
            cs = core_wb[name]
            es = enr_wb[name]
            assert cs.max_row == es.max_row
            assert cs.max_column == es.max_column
            for row in cs.iter_rows(values_only=True):
                pass
            for r in range(1, cs.max_row + 1):
                for c in range(1, cs.max_column + 1):
                    assert cs.cell(r, c).value == es.cell(r, c).value


@pytest.mark.asyncio
async def test_complete_enrichment_idempotent_with_old_token(
    stage3_session_maker: async_sessionmaker[AsyncSession],
    valid_result: ValidationResult,
) -> None:
    async with stage3_session_maker() as session:
        cycle_id = await complete_cycle(
            session,
            valid_result,
            report_date=dt.date(2026, 11, 4),
            sha_prefix="s44-idemp",
        )
        await session.commit()
    report_id, _, _ = await build_ready_core(stage3_session_maker, cycle_id=cycle_id)
    settings = _settings()
    async with stage3_session_maker() as session:
        pos = await session.scalar(select(DebtPosition).limit(1))
        assert pos is not None
        pos.comment_raw = "Оплата 21.11"
        await session.commit()
    async with stage3_session_maker() as session:
        job_id = await enqueue_enrichment_job(
            session, report_id=report_id, settings=settings
        )
    async with stage3_session_maker() as session:
        claim = await claim_enrichment_job(
            session, job_id=job_id, settings=settings
        )
    assert claim is not None
    first = await run_claimed_enrichment(
        stage3_session_maker, claim=claim, settings=settings, llm_client=None
    )
    assert first is not None
    # Retry complete with old token — idempotent success
    async with stage3_session_maker() as session:
        again = await complete_enrichment(
            session,
            job_id=claim.job_id,
            claim_token=claim.claim_token,
            excel_bytes=b"ignored",
            excel_sha256="0" * 64,
            enrichment_counts={},
        )
    assert again.idempotent is True
    assert again.artifact_id == first.artifact_id
    async with stage3_session_maker() as session:
        count = await session.scalar(
            select(func.count()).select_from(AuditArtifact).where(
                AuditArtifact.audit_report_id == report_id,
                AuditArtifact.kind == AuditArtifactKind.ENRICHED,
            )
        )
        assert count == 1
        deliveries = await session.scalar(
            select(func.count()).select_from(ReportDelivery).where(
                ReportDelivery.kind == ReportDeliveryKind.AUTOMATIC,
            )
        )
        # CORE + ENRICHED automatic
        assert deliveries == 2


@pytest.mark.asyncio
async def test_record_analysis_immutable_and_terminal_retry(
    stage3_session_maker: async_sessionmaker[AsyncSession],
    valid_result: ValidationResult,
) -> None:
    async with stage3_session_maker() as session:
        cycle_id = await complete_cycle(
            session,
            valid_result,
            report_date=dt.date(2026, 11, 5),
            sha_prefix="s44-imm",
        )
        await session.commit()
    report_id, _, _ = await build_ready_core(stage3_session_maker, cycle_id=cycle_id)
    settings = _settings(comment_enrichment_max_attempts=1)
    async with stage3_session_maker() as session:
        pos = await session.scalar(select(DebtPosition).limit(1))
        assert pos is not None
        position_id = int(pos.id)
        pos.comment_raw = "непонятный текст без фактов"
        await session.commit()
    async with stage3_session_maker() as session:
        job_id = await enqueue_enrichment_job(
            session, report_id=report_id, settings=settings
        )
    async with stage3_session_maker() as session:
        claim = await claim_enrichment_job(
            session, job_id=job_id, settings=settings
        )
    assert claim is not None

    class BoomClient:
        async def analyze_comment(self, **kwargs: object) -> LlmCommentFacts:
            raise OpenRouterTransientError("boom")

    # With API key path via injectable that fails transiently
    settings2 = _settings(
        comment_enrichment_max_attempts=1,
        openrouter_api_key="x",
        openrouter_model="m",
    )
    await run_claimed_enrichment(
        stage3_session_maker,
        claim=claim,
        settings=settings2,
        llm_client=BoomClient(),  # type: ignore[arg-type]
    )
    async with stage3_session_maker() as session:
        job = await session.get(CommentEnrichmentJob, job_id)
        assert job is not None
        assert job.status == CommentEnrichmentJobStatus.FAILED
        assert job.next_retry_at is None
        assert await session.scalar(
            select(AuditArtifact.id).where(AuditArtifact.enrichment_job_id == job_id)
        ) is None

    async with stage3_session_maker() as session:
        assert await retry_terminal_enrichment(session, job_id=job_id) is True
        job = await session.get(CommentEnrichmentJob, job_id)
        assert job is not None
        assert job.status == CommentEnrichmentJobStatus.PENDING
        assert job.attempt_count == 0
        assert job.operator_retry_count == 1

    # Immutable record_comment_analysis
    async with stage3_session_maker() as session:
        claim2 = await claim_enrichment_job(
            session, job_id=job_id, settings=settings2
        )
    assert claim2 is not None
    async with stage3_session_maker() as session:
        ok = await record_comment_analysis(
            session,
            job_id=job_id,
            claim_token=claim2.claim_token,
            debt_position_id=position_id,
            analysis_input_hash="h1",
            comment_raw="непонятный текст без фактов",
            source=CommentAnalysisSource.UNPARSED,
            analysis_status=CommentAnalysisStatus.NEEDS_REVIEW,
            confidence=CommentAnalysisConfidence.NONE,
            summary="x",
        )
        assert ok is True
    async with stage3_session_maker() as session:
        ok2 = await record_comment_analysis(
            session,
            job_id=job_id,
            claim_token=claim2.claim_token,
            debt_position_id=position_id,
            analysis_input_hash="h1",
            comment_raw="непонятный текст без фактов",
            source=CommentAnalysisSource.UNPARSED,
            analysis_status=CommentAnalysisStatus.NEEDS_REVIEW,
            confidence=CommentAnalysisConfidence.NONE,
            summary="x",
        )
        assert ok2 is True
    async with stage3_session_maker() as session:
        with pytest.raises(EnrichmentConflictError):
            await record_comment_analysis(
                session,
                job_id=job_id,
                claim_token=claim2.claim_token,
                debt_position_id=position_id,
                analysis_input_hash="h2",
                comment_raw="другое",
                source=CommentAnalysisSource.UNPARSED,
                analysis_status=CommentAnalysisStatus.NEEDS_REVIEW,
                confidence=CommentAnalysisConfidence.NONE,
                summary="y",
            )
    async with stage3_session_maker() as session:
        count = await session.scalar(
            select(func.count()).select_from(CommentAnalysis).where(
                CommentAnalysis.enrichment_job_id == job_id
            )
        )
        assert count == 1


@pytest.mark.asyncio
async def test_scheduler_delivers_core_before_hung_enrichment(
    stage3_session_maker: async_sessionmaker[AsyncSession],
    valid_result: ValidationResult,
) -> None:
    async with stage3_session_maker() as session:
        cycle_id = await complete_cycle(
            session,
            valid_result,
            report_date=dt.date(2026, 11, 6),
            sha_prefix="s44-ord",
        )
        await session.commit()
    report_id, _, _ = await build_ready_core(stage3_session_maker, cycle_id=cycle_id)
    settings = _settings(
        openrouter_api_key="x",
        openrouter_model="m",
        comment_enrichment_run_timeout_seconds=30,
    )
    async with stage3_session_maker() as session:
        pos = await session.scalar(select(DebtPosition).limit(1))
        assert pos is not None
        pos.comment_raw = "уточнить детали"
        await session.commit()
    async with stage3_session_maker() as session:
        await enqueue_enrichment_job(
            session, report_id=report_id, settings=settings
        )

    sent_docs: list[int] = []
    hang_started = False

    class HangClient:
        async def analyze_comment(self, **kwargs: object) -> LlmCommentFacts:
            nonlocal hang_started
            hang_started = True
            raise OpenRouterTransientError("hang")

    async def fake_send_message(*, chat_id: int, text: str, **kwargs: object) -> object:
        return SimpleNamespace(message_id=1)

    async def fake_send_document(*, chat_id: int, document: object, **kwargs: object) -> object:
        sent_docs.append(chat_id)
        # Enrichment must not have started before CORE document send in this tick
        assert hang_started is False
        return SimpleNamespace(message_id=2)

    scheduler = ReportScheduler(
        bot=AsyncMock(),
        session_maker=stage3_session_maker,
        settings=settings,  # type: ignore[arg-type]
        send_message=fake_send_message,
        send_document=fake_send_document,
        llm_client=HangClient(),  # type: ignore[arg-type]
    )
    await scheduler.run_once()
    assert sent_docs == [NOTIFY_CHAT_ID]
    async with stage3_session_maker() as session:
        core_delivery = await session.scalar(
            select(ReportDelivery)
            .join(AuditArtifact, AuditArtifact.id == ReportDelivery.audit_artifact_id)
            .where(
                AuditArtifact.kind == AuditArtifactKind.CORE,
                ReportDelivery.status == ReportDeliveryStatus.DELIVERED,
            )
        )
        assert core_delivery is not None
