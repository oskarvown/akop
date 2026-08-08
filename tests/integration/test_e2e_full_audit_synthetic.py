"""Сквозной E2E-аудит: 6 циклов × 5 отделов на реальных и синтетических Excel.

Цикл 1 — оригиналы из Downloads (30.07.2026); циклы 2–6 — synthetic_e2e.
Проверяет add, подтверждение отдела, replace, /undo, auto-complete 5/5,
неизменяемость COMPLETED и сверку с manifest.

Изоляция данных: использует фикстуру ``stage3_session_maker`` из
``tests/integration/conftest.py`` — перед и после теста полностью очищаются
таблицы Stage 3 в PostgreSQL. Cleanup разрешён только для БД ``*_test`` или при
``ALLOW_DESTRUCTIVE_TEST_DB=1`` (см. ``tests/integration/db_safety.py``).

Запуск: ``pytest -m e2e tests/integration/test_e2e_full_audit_synthetic.py``
"""
from __future__ import annotations

import csv
import datetime as dt
import shutil
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.application.audit_service import find_audit_cycle_by_report_date
from app.bot.handlers.status import handle_status
from app.bot.handlers.upload import (
    UploadStates,
    handle_department_callback,
    handle_new_cycle_callback,
    handle_replace_callback,
    handle_undo,
    handle_undo_confirm_callback,
)
from app.domain.enums import Department
from app.domain.models import AuditCycle, AuditCycleStatus, SourceFile, SourceFileLifecycle
from app.infrastructure.excel.validator import validate_confirmed_template_file
from tests.integration.test_bot_upload_flow import (
    FakeCallback,
    FakeMessage,
    FakeState,
    choose_and_confirm_department,
    choose_department,
    confirm_department,
    receive_valid,
)

DOWNLOADS = Path("/Users/efraimdiverolli/Downloads")
SYNTHETIC = Path(__file__).resolve().parents[2] / "private_inputs" / "synthetic_e2e"
MANIFEST_SUMMARY = SYNTHETIC / "manifest" / "summary_by_cycle.csv"

CYCLE_SPECS: tuple[tuple[dt.date, Path], ...] = (
    (dt.date(2026, 7, 30), DOWNLOADS),
    (dt.date(2026, 8, 6), SYNTHETIC / "cycle_2026-08-06"),
    (dt.date(2026, 8, 13), SYNTHETIC / "cycle_2026-08-13"),
    (dt.date(2026, 8, 20), SYNTHETIC / "cycle_2026-08-20"),
    (dt.date(2026, 8, 27), SYNTHETIC / "cycle_2026-08-27"),
    (dt.date(2026, 9, 3), SYNTHETIC / "cycle_2026-09-03"),
)

DEPT_FILE_TPL: dict[Department, str] = {
    Department.SZFO_1: "Дебиторка_СЗФО-1_{d}.xlsx",
    Department.SZFO_2: "Дебиторка_СЗФО-2_{d}.xlsx",
    Department.REGIONAL: "Дебиторка_региональный_отдел_{d}.xlsx",
    Department.MOSCOW: "Дебиторка_Москва_{d}.xlsx",
    Department.FOKIN: "Дебиторка_Фокин_{d}.xlsx",
}

# Оригинал регионального файла в Downloads имеет дату в имени через underscore.
DOWNLOADS_REGIONAL = DOWNLOADS / "Дебиторка_региональный_отдел_30_07.xlsx"

BASELINE_DOWNLOADS: tuple[Path, ...] = (
    DOWNLOADS / "Дебиторка_СЗФО-1_30.07.2026.xlsx",
    DOWNLOADS / "Дебиторка_СЗФО-2_30.07.2026.xlsx",
    DOWNLOADS_REGIONAL,
    DOWNLOADS / "Дебиторка_Москва_30.07.2026.xlsx",
    DOWNLOADS / "Дебиторка_Фокин_30.07.2026.xlsx",
)


def _require_e2e_fixtures() -> None:
    """Skip с понятной причиной, если локальные Excel-fixtures недоступны."""
    missing: list[str] = []
    if not DOWNLOADS.is_dir():
        missing.append(f"каталог Downloads отсутствует: {DOWNLOADS}")
    else:
        for path in BASELINE_DOWNLOADS:
            if not path.is_file():
                missing.append(f"оригинал: {path}")
    if not SYNTHETIC.is_dir():
        missing.append(f"synthetic_e2e: каталог отсутствует ({SYNTHETIC})")
    elif not MANIFEST_SUMMARY.is_file():
        missing.append(f"synthetic_e2e: нет manifest ({MANIFEST_SUMMARY})")
    else:
        for report_date, base in CYCLE_SPECS[1:]:
            if not base.is_dir():
                missing.append(f"synthetic_e2e: нет {base}")
                continue
            for dept in Department:
                path = dept_path(base, dept, report_date)
                if not path.is_file():
                    missing.append(f"synthetic_e2e: нет {path}")
    if missing:
        pytest.skip(
            "E2E-fixtures недоступны (нужны оригиналы в ~/Downloads и "
            f"private_inputs/synthetic_e2e):\n- " + "\n- ".join(missing)
        )


def _fmt(d: dt.date) -> str:
    return d.strftime("%d.%m.%Y")


def dept_path(base: Path, department: Department, report_date: dt.date) -> Path:
    if base == DOWNLOADS and department == Department.REGIONAL and report_date == dt.date(2026, 7, 30):
        return DOWNLOADS_REGIONAL
    return base / DEPT_FILE_TPL[department].format(d=_fmt(report_date))


def expected_debt(path: Path) -> Decimal:
    result = validate_confirmed_template_file(path)
    assert result.is_valid and result.parsed is not None
    assert result.parsed.grand_total.total_debt is not None
    return result.parsed.grand_total.total_debt


def touched_copy(src: Path, dst: Path, *, row: int = 11, col: int = 9, delta: float = -1.0) -> Path:
    """Копия файла с другим SHA, та же report_date — для replace/immutability."""
    shutil.copy2(src, dst)
    wb = openpyxl.load_workbook(dst)
    ws = wb["TDSheet"]
    current = ws.cell(row, col).value or 0
    ws.cell(row, col).value = float(current) + delta
    wb.save(dst)
    wb.close()
    return dst


async def _active_debt(session: AsyncSession, report_date: dt.date, department: Department) -> Decimal:
    async with session.begin():
        row = await session.scalar(
            select(SourceFile.reported_grand_totals).where(
                SourceFile.department == department,
                SourceFile.lifecycle_status == SourceFileLifecycle.ACTIVE,
                SourceFile.report_date == report_date,
            )
        )
    assert row is not None
    return Decimal(str(row["total_debt"]))


async def upload_and_confirm(
    session: AsyncSession,
    state: FakeState,
    path: Path,
    department: Department,
) -> FakeCallback:
    await receive_valid(session, state, path)
    await choose_department(session, state, department)
    if state.state == UploadStates.confirming_new_cycle.state:
        token = state.data["upload_token"]
        await handle_new_cycle_callback(
            FakeCallback(f"newcycle:{token}:confirm"), state, session  # type: ignore[arg-type]
        )
        await choose_department(session, state, department)
    assert state.state == UploadStates.confirming_department.state
    return await confirm_department(session, state)


@pytest.mark.integration
@pytest.mark.e2e
@pytest.mark.asyncio
async def test_full_six_cycle_e2e_audit_with_real_and_synthetic_files(
    stage3_session: AsyncSession,
    tmp_path: Path,
) -> None:
    _require_e2e_fixtures()

    state = FakeState()
    baseline_date = CYCLE_SPECS[0][0]
    baseline_base = CYCLE_SPECS[0][1]

    # --- Цикл 1: add + undo (неверный отдел) + replace + complete ---
    await upload_and_confirm(
        stage3_session, state, dept_path(baseline_base, Department.SZFO_1, baseline_date), Department.SZFO_1
    )
    await upload_and_confirm(
        stage3_session, state, dept_path(baseline_base, Department.SZFO_2, baseline_date), Department.SZFO_2
    )

    # Undo: файл Москвы назначаем ошибочно на Региональный, затем отменяем.
    moscow_path = dept_path(baseline_base, Department.MOSCOW, baseline_date)
    await receive_valid(stage3_session, state, moscow_path)
    await choose_department(stage3_session, state, Department.REGIONAL)
    await confirm_department(stage3_session, state)
    await handle_undo(FakeMessage(), state)  # type: ignore[arg-type]
    undo_cb = FakeCallback(f"undoconfirm:{state.data['undo_token']}:confirm")
    await handle_undo_confirm_callback(undo_cb, state, stage3_session)  # type: ignore[arg-type]
    assert "Снял файл" in undo_cb.message.answers[-1][0]
    assert state.state == UploadStates.choosing_department.state

    await choose_and_confirm_department(stage3_session, state, Department.MOSCOW)

    await upload_and_confirm(
        stage3_session, state, dept_path(baseline_base, Department.REGIONAL, baseline_date), Department.REGIONAL
    )

    # Replace: вариант регионального с той же датой, другой SHA.
    regional_src = dept_path(baseline_base, Department.REGIONAL, baseline_date)
    replace_path = touched_copy(regional_src, tmp_path / "regional_replace_30.07.xlsx")

    await receive_valid(stage3_session, state, replace_path)
    await choose_and_confirm_department(stage3_session, state, Department.REGIONAL)
    assert state.state == UploadStates.confirming_replace.state
    rep_cb = FakeCallback(f"replace:{state.data['upload_token']}:confirm")
    await handle_replace_callback(rep_cb, state, stage3_session)  # type: ignore[arg-type]
    assert "4/5" in rep_cb.message.answers[-1][0]

    fokin_cb = await upload_and_confirm(
        stage3_session, state, dept_path(baseline_base, Department.FOKIN, baseline_date), Department.FOKIN
    )
    assert "5/5" in fokin_cb.message.answers[-1][0]
    assert state.state is None

    async with stage3_session.begin():
        cycle1 = await stage3_session.scalar(
            select(AuditCycle).where(AuditCycle.report_date == baseline_date)
        )
        assert cycle1 is not None
        assert cycle1.status == AuditCycleStatus.COMPLETED
        active_count = await stage3_session.scalar(
            select(func.count(SourceFile.id)).where(
                SourceFile.audit_cycle_id == cycle1.id,
                SourceFile.lifecycle_status == SourceFileLifecycle.ACTIVE,
            )
        )
        assert active_count == 5
        superseded = await stage3_session.scalar(
            select(func.count(SourceFile.id)).where(
                SourceFile.lifecycle_status == SourceFileLifecycle.SUPERSEDED,
            )
        )
        assert superseded == 1
        revoked = await stage3_session.scalar(
            select(func.count(SourceFile.id)).where(
                SourceFile.lifecycle_status == SourceFileLifecycle.REVOKED,
            )
        )
        assert revoked == 1

    # Сверка baseline с парсером (manifest для 30.07 не генерировался).
    for dept in Department:
        path = dept_path(baseline_base, dept, baseline_date)
        if dept == Department.REGIONAL:
            path = replace_path  # active после replace
        exp = expected_debt(path if dept != Department.REGIONAL else replace_path)
        if dept == Department.REGIONAL:
            got = await _active_debt(stage3_session, baseline_date, dept)
            assert abs(got - exp) <= Decimal("0.01")
        else:
            got = await _active_debt(stage3_session, baseline_date, dept)
            assert abs(got - expected_debt(dept_path(baseline_base, dept, baseline_date))) <= Decimal("0.01")

    # COMPLETED цикл 1 неизменяем (новый SHA, та же дата).
    state_block = FakeState()
    immut_path = touched_copy(
        dept_path(baseline_base, Department.SZFO_1, baseline_date),
        tmp_path / "szfo1_immutability_probe.xlsx",
    )
    await receive_valid(stage3_session, state_block, immut_path)
    token = state_block.data["upload_token"]
    block_cb = FakeCallback(f"dept:{token}:{Department.SZFO_1.value}")
    await handle_department_callback(block_cb, state_block, stage3_session)  # type: ignore[arg-type]
    assert "завершён" in block_cb.message.answers[-1][0]

    # --- Циклы 2–6: полные комплекты, auto-complete ---
    manifest: dict[tuple[str, str], Decimal] = {}
    with MANIFEST_SUMMARY.open(encoding="utf-8") as f:
        for row in csv.DictReader(f):
            manifest[(row["cycle_date"], row["department"])] = Decimal(row["grand_total_debt_after"])

    for report_date, base in CYCLE_SPECS[1:]:
        for dept in Department:
            path = dept_path(base, dept, report_date)
            cb = await upload_and_confirm(stage3_session, state, path, dept)
            if dept == Department.FOKIN:
                assert "5/5" in cb.message.answers[-1][0]

        async with stage3_session.begin():
            cycle = await stage3_session.scalar(
                select(AuditCycle).where(AuditCycle.report_date == report_date)
            )
            assert cycle is not None
            assert cycle.status == AuditCycleStatus.COMPLETED

        for dept in Department:
            exp = manifest[(_fmt(report_date), dept.value)]
            got = await _active_debt(stage3_session, report_date, dept)
            assert abs(got - exp) <= Decimal("0.01"), f"{dept.value} {_fmt(report_date)}: {got} != {exp}"

    async with stage3_session.begin():
        completed = await stage3_session.scalar(
            select(func.count(AuditCycle.id)).where(
                AuditCycle.status == AuditCycleStatus.COMPLETED
            )
        )
        assert completed == 6

    status_msg = FakeMessage()
    await handle_status(status_msg, stage3_session)  # type: ignore[arg-type]
    status_text = "\n".join(text for text, _ in status_msg.answers)
    # /status показывает collecting + 3 новейших terminal-цикла (см. list_cycle_statuses).
    assert "03.09.2026" in status_text
    assert status_text.count("5/5") >= 3
    assert "06.08.2026" not in status_text

    # Все COMPLETED циклы неизменяем (проба первого и последнего с новым SHA).
    for report_date, base in (CYCLE_SPECS[0], CYCLE_SPECS[-1]):
        st = FakeState()
        probe = touched_copy(
            dept_path(base, Department.FOKIN, report_date),
            tmp_path / f"fokin_immut_{report_date.isoformat()}.xlsx",
        )
        await receive_valid(stage3_session, st, probe)
        t = st.data["upload_token"]
        cb = FakeCallback(f"dept:{t}:{Department.FOKIN.value}")
        await handle_department_callback(cb, st, stage3_session)  # type: ignore[arg-type]
        assert "завершён" in cb.message.answers[-1][0]
        assert await find_audit_cycle_by_report_date(stage3_session, report_date) is not None
