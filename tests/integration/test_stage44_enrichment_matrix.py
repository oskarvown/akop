"""Stage 4.4 locked matrix coverage beyond the smoke suite."""
from __future__ import annotations

import asyncio
import datetime as dt
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import AsyncMock

import pytest
from openpyxl import load_workbook
from sqlalchemy import func, select, text
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker

from app.application.comment_enrichment_service import (
    claim_enrichment_job,
    compute_analysis_input_hash,
    enqueue_enrichment_job,
    fail_enrichment,
    record_comment_analysis,
    run_claimed_enrichment,
)
from app.application.report_delivery_service import (
    claim_delivery,
    create_manual_delivery,
    load_delivery_send_context,
    record_document_sent,
    recover_missing_automatic_deliveries,
)
from app.bot.scheduler.report_scheduler import ReportScheduler
from app.domain.models import (
    AuditArtifact,
    AuditArtifactKind,
    AuditReport,
    CommentAnalysis,
    CommentAnalysisConfidence,
    CommentAnalysisSource,
    CommentAnalysisStatus,
    CommentEnrichmentJob,
    CommentEnrichmentJobStatus,
    DebtPosition,
    ReportDelivery,
    ReportDeliveryKind,
    ReportDeliveryStatus,
)
from app.infrastructure.excel.validator import (
    ValidationResult,
    validate_confirmed_template_file,
)
from app.infrastructure.llm.openrouter_client import (
    LlmCommentFacts,
    OpenRouterSchemaError,
    OpenRouterTransientError,
)
from tests.integration.test_stage44_enrichment import (
    NOTIFY_CHAT_ID,
    _settings,
    build_ready_core,
    complete_cycle,
)

FIXTURE = (
    Path(__file__).resolve().parents[1]
    / "fixtures"
    / "regional"
    / "regional_valid_basic.xls"
)


@pytest.fixture
def valid_result() -> ValidationResult:
    result = validate_confirmed_template_file(FIXTURE)
    assert result.is_valid and result.parsed is not None
    return result


async def _set_only_comment(
    maker: async_sessionmaker[AsyncSession], *, comment: str
) -> int:
    async with maker() as session:
        await session.execute(text("UPDATE debt_positions SET comment_raw = NULL"))
        pos = await session.scalar(select(DebtPosition).order_by(DebtPosition.id).limit(1))
        assert pos is not None
        pos.comment_raw = comment
        position_id = int(pos.id)
        await session.commit()
        return position_id


async def _set_two_comments(
    maker: async_sessionmaker[AsyncSession], *, first: str, second: str
) -> tuple[int, int]:
    async with maker() as session:
        await session.execute(text("UPDATE debt_positions SET comment_raw = NULL"))
        positions = (
            await session.execute(select(DebtPosition).order_by(DebtPosition.id).limit(2))
        ).scalars().all()
        assert len(positions) >= 2
        positions[0].comment_raw = first
        positions[1].comment_raw = second
        ids = (int(positions[0].id), int(positions[1].id))
        await session.commit()
        return ids



@pytest.mark.asyncio
async def test_concurrent_identical_hash_enqueue_one_job(
    stage3_session_maker: async_sessionmaker[AsyncSession],
    valid_result: ValidationResult,
) -> None:
    async with stage3_session_maker() as session:
        cycle_id = await complete_cycle(
            session, valid_result, report_date=dt.date(2026, 12, 1), sha_prefix="s44c-hash"
        )
        await session.commit()
    report_id, _, _ = await build_ready_core(stage3_session_maker, cycle_id=cycle_id)
    settings = _settings()
    async with stage3_session_maker() as session:
        pos = await session.scalar(select(DebtPosition).limit(1))
        assert pos is not None
        pos.comment_raw = "Оплата 01.12"
        await session.commit()

    async def _enqueue() -> int | None:
        async with stage3_session_maker() as session:
            return await enqueue_enrichment_job(
                session, report_id=report_id, settings=settings
            )

    ids = await asyncio.gather(_enqueue(), _enqueue(), _enqueue())
    assert len(set(ids)) == 1
    assert ids[0] is not None
    async with stage3_session_maker() as session:
        count = await session.scalar(
            select(func.count()).select_from(CommentEnrichmentJob).where(
                CommentEnrichmentJob.audit_report_id == report_id
            )
        )
        assert count == 1


@pytest.mark.asyncio
async def test_concurrent_revision_assignment_unique(
    stage3_session_maker: async_sessionmaker[AsyncSession],
    valid_result: ValidationResult,
) -> None:
    async with stage3_session_maker() as session:
        cycle_id = await complete_cycle(
            session, valid_result, report_date=dt.date(2026, 12, 2), sha_prefix="s44c-rev"
        )
        await session.commit()
    report_id, _, _ = await build_ready_core(stage3_session_maker, cycle_id=cycle_id)
    async with stage3_session_maker() as session:
        pos = await session.scalar(select(DebtPosition).limit(1))
        assert pos is not None
        pos.comment_raw = "Оплата 02.12"
        await session.commit()

    settings_a = _settings(openrouter_model="m-a")
    settings_b = _settings(openrouter_model="m-b")
    async with stage3_session_maker() as session:
        job_a = await enqueue_enrichment_job(
            session, report_id=report_id, settings=settings_a
        )
        job_b = await enqueue_enrichment_job(
            session, report_id=report_id, settings=settings_b
        )
    assert job_a is not None and job_b is not None and job_a != job_b

    async with stage3_session_maker() as session:
        claim_a = await claim_enrichment_job(
            session, job_id=job_a, settings=settings_a
        )
        claim_b = await claim_enrichment_job(
            session, job_id=job_b, settings=settings_b
        )
    assert claim_a and claim_b

    results = await asyncio.gather(
        run_claimed_enrichment(
            stage3_session_maker, claim=claim_a, settings=settings_a, llm_client=None
        ),
        run_claimed_enrichment(
            stage3_session_maker, claim=claim_b, settings=settings_b, llm_client=None
        ),
    )
    revisions = sorted(r.revision for r in results if r is not None)
    assert revisions == [1, 2]
    async with stage3_session_maker() as session:
        arts = (
            await session.execute(
                select(AuditArtifact.revision).where(
                    AuditArtifact.audit_report_id == report_id,
                    AuditArtifact.kind == AuditArtifactKind.ENRICHED,
                )
            )
        ).scalars().all()
        assert sorted(arts) == [1, 2]


@pytest.mark.asyncio
async def test_r1_r2_analyses_immutable_and_isolated(
    stage3_session_maker: async_sessionmaker[AsyncSession],
    valid_result: ValidationResult,
) -> None:
    async with stage3_session_maker() as session:
        cycle_id = await complete_cycle(
            session, valid_result, report_date=dt.date(2026, 12, 3), sha_prefix="s44c-r12"
        )
        await session.commit()
    report_id, _, _ = await build_ready_core(stage3_session_maker, cycle_id=cycle_id)
    async with stage3_session_maker() as session:
        pos = await session.scalar(select(DebtPosition).limit(1))
        assert pos is not None
        position_id = int(pos.id)
        pos.comment_raw = "Оплата 03.12 на 1000"
        await session.commit()

    settings_r1 = _settings(openrouter_model="m-r1")
    async with stage3_session_maker() as session:
        job1 = await enqueue_enrichment_job(
            session, report_id=report_id, settings=settings_r1
        )
    async with stage3_session_maker() as session:
        claim1 = await claim_enrichment_job(
            session, job_id=job1, settings=settings_r1
        )
    assert claim1 is not None
    r1 = await run_claimed_enrichment(
        stage3_session_maker, claim=claim1, settings=settings_r1, llm_client=None
    )
    assert r1 is not None and r1.revision == 1

    settings_r2 = _settings(openrouter_model="m-r2")
    async with stage3_session_maker() as session:
        job2 = await enqueue_enrichment_job(
            session, report_id=report_id, settings=settings_r2
        )
    async with stage3_session_maker() as session:
        claim2 = await claim_enrichment_job(
            session, job_id=job2, settings=settings_r2
        )
    assert claim2 is not None
    r2 = await run_claimed_enrichment(
        stage3_session_maker, claim=claim2, settings=settings_r2, llm_client=None
    )
    assert r2 is not None and r2.revision == 2

    async with stage3_session_maker() as session:
        a1 = await session.scalar(
            select(CommentAnalysis).where(
                CommentAnalysis.enrichment_job_id == job1,
                CommentAnalysis.debt_position_id == position_id,
            )
        )
        a2 = await session.scalar(
            select(CommentAnalysis).where(
                CommentAnalysis.enrichment_job_id == job2,
                CommentAnalysis.debt_position_id == position_id,
            )
        )
        assert a1 is not None and a2 is not None
        assert a1.id != a2.id
        assert a1.analysis_status is CommentAnalysisStatus.RESOLVED


@pytest.mark.asyncio
async def test_old_token_cannot_progress_or_fail(
    stage3_session_maker: async_sessionmaker[AsyncSession],
    valid_result: ValidationResult,
) -> None:
    async with stage3_session_maker() as session:
        cycle_id = await complete_cycle(
            session, valid_result, report_date=dt.date(2026, 12, 4), sha_prefix="s44c-fence"
        )
        await session.commit()
    report_id, _, _ = await build_ready_core(stage3_session_maker, cycle_id=cycle_id)
    settings = _settings()
    async with stage3_session_maker() as session:
        pos = await session.scalar(select(DebtPosition).limit(1))
        assert pos is not None
        position_id = int(pos.id)
        pos.comment_raw = "уточнить детали"
        await session.commit()
    async with stage3_session_maker() as session:
        job_id = await enqueue_enrichment_job(
            session, report_id=report_id, settings=settings
        )
    async with stage3_session_maker() as session:
        claim = await claim_enrichment_job(
            session, job_id=job_id, settings=settings
        )
    assert claim is not None
    old_token = claim.claim_token
    # Force reclaim path: mark failed and claim again
    async with stage3_session_maker() as session:
        job = await session.get(CommentEnrichmentJob, job_id)
        assert job is not None
        job.status = CommentEnrichmentJobStatus.FAILED
        job.claim_token = None
        job.claimed_at = None
        job.next_retry_at = dt.datetime.now(dt.timezone.utc) - dt.timedelta(seconds=1)
        await session.commit()
    async with stage3_session_maker() as session:
        claim2 = await claim_enrichment_job(
            session, job_id=job_id, settings=settings
        )
    assert claim2 is not None
    assert claim2.claim_token != old_token

    async with stage3_session_maker() as session:
        assert (
            await record_comment_analysis(
                session,
                job_id=job_id,
                claim_token=old_token,
                debt_position_id=position_id,
                analysis_input_hash="h",
                comment_raw="уточнить детали",
                source=CommentAnalysisSource.UNPARSED,
                analysis_status=CommentAnalysisStatus.NEEDS_REVIEW,
                confidence=CommentAnalysisConfidence.NONE,
                summary="x",
            )
            is False
        )
        assert (
            await fail_enrichment(
                session,
                job_id=job_id,
                claim_token=old_token,
                error="stale",
                settings=settings,
            )
            is False
        )
        job = await session.get(CommentEnrichmentJob, job_id)
        assert job is not None
        assert job.status is CommentEnrichmentJobStatus.CLAIMED


@pytest.mark.asyncio
async def test_run_timeout_before_claim_ttl(
    stage3_session_maker: async_sessionmaker[AsyncSession],
    valid_result: ValidationResult,
) -> None:
    async with stage3_session_maker() as session:
        cycle_id = await complete_cycle(
            session, valid_result, report_date=dt.date(2026, 12, 5), sha_prefix="s44c-tto"
        )
        await session.commit()
    report_id, _, _ = await build_ready_core(stage3_session_maker, cycle_id=cycle_id)
    settings = _settings(
        comment_enrichment_claim_ttl_seconds=60,
        comment_enrichment_run_timeout_seconds=0.05,
        openrouter_api_key="k",
        openrouter_model="m",
    )
    assert (
        settings.comment_enrichment_run_timeout_seconds
        < settings.comment_enrichment_claim_ttl_seconds
    )
    async with stage3_session_maker() as session:
        pos = await session.scalar(select(DebtPosition).limit(1))
        assert pos is not None
        pos.comment_raw = "уточнить позже"
        await session.commit()
    async with stage3_session_maker() as session:
        job_id = await enqueue_enrichment_job(
            session, report_id=report_id, settings=settings
        )
    async with stage3_session_maker() as session:
        claim = await claim_enrichment_job(
            session, job_id=job_id, settings=settings
        )
    assert claim is not None

    hang = asyncio.Event()

    class HangClient:
        async def analyze_comment(self, **kwargs: object) -> LlmCommentFacts:
            await hang.wait()
            raise AssertionError("should timeout")

    await run_claimed_enrichment(
        stage3_session_maker,
        claim=claim,
        settings=settings,
        llm_client=HangClient(),  # type: ignore[arg-type]
    )
    async with stage3_session_maker() as session:
        job = await session.get(CommentEnrichmentJob, job_id)
        assert job is not None
        assert job.status is CommentEnrichmentJobStatus.FAILED
        assert job.last_error == "enrichment_run_timeout"
        assert job.status is not CommentEnrichmentJobStatus.CLAIMED


@pytest.mark.asyncio
async def test_transient_resume_skips_saved_without_rellm(
    stage3_session_maker: async_sessionmaker[AsyncSession],
    valid_result: ValidationResult,
) -> None:
    async with stage3_session_maker() as session:
        cycle_id = await complete_cycle(
            session, valid_result, report_date=dt.date(2026, 12, 6), sha_prefix="s44c-res"
        )
        await session.commit()
    report_id, _, _ = await build_ready_core(stage3_session_maker, cycle_id=cycle_id)
    settings = _settings(
        openrouter_api_key="k",
        openrouter_model="m",
        comment_enrichment_max_attempts=3,
    )
    _, amb_id = await _set_two_comments(
        stage3_session_maker,
        first="Оплата 06.12 на 500",
        second="уточнить детали",
    )

    async with stage3_session_maker() as session:
        job_id = await enqueue_enrichment_job(
            session, report_id=report_id, settings=settings
        )
    async with stage3_session_maker() as session:
        claim = await claim_enrichment_job(
            session, job_id=job_id, settings=settings
        )
    assert claim is not None

    calls: list[int] = []

    class BoomThenOk:
        async def analyze_comment(self, **kwargs: object) -> LlmCommentFacts:
            calls.append(1)
            if len(calls) == 1:
                raise OpenRouterTransientError("once")
            return LlmCommentFacts(
                mentioned_date=None,
                mentioned_amount=None,
                action=None,
                reason=None,
                responsible_person=None,
                summary="llm",
                confidence="low",
                raw_json={"confidence": "low"},
            )

    await run_claimed_enrichment(
        stage3_session_maker,
        claim=claim,
        settings=settings,
        llm_client=BoomThenOk(),  # type: ignore[arg-type]
    )
    async with stage3_session_maker() as session:
        job = await session.get(CommentEnrichmentJob, job_id)
        assert job is not None
        assert job.status is CommentEnrichmentJobStatus.FAILED
        saved = await session.scalar(
            select(func.count()).select_from(CommentAnalysis).where(
                CommentAnalysis.enrichment_job_id == job_id,
                CommentAnalysis.analysis_status == CommentAnalysisStatus.RESOLVED,
            )
        )
        assert saved == 1
        job.next_retry_at = dt.datetime.now(dt.timezone.utc) - dt.timedelta(seconds=1)
        await session.commit()

    async with stage3_session_maker() as session:
        claim2 = await claim_enrichment_job(
            session, job_id=job_id, settings=settings
        )
    assert claim2 is not None
    await run_claimed_enrichment(
        stage3_session_maker,
        claim=claim2,
        settings=settings,
        llm_client=BoomThenOk(),  # type: ignore[arg-type]
    )
    # Second run: only the unfinished ambiguous comment should call LLM once more
    assert len(calls) == 2
    async with stage3_session_maker() as session:
        job = await session.get(CommentEnrichmentJob, job_id)
        assert job is not None
        assert job.status is CommentEnrichmentJobStatus.READY
        amb = await session.scalar(
            select(CommentAnalysis).where(
                CommentAnalysis.enrichment_job_id == job_id,
                CommentAnalysis.debt_position_id == amb_id,
            )
        )
        assert amb is not None
        assert amb.source is CommentAnalysisSource.LLM


@pytest.mark.asyncio
async def test_schema_invalid_needs_review_not_transient(
    stage3_session_maker: async_sessionmaker[AsyncSession],
    valid_result: ValidationResult,
) -> None:
    async with stage3_session_maker() as session:
        cycle_id = await complete_cycle(
            session, valid_result, report_date=dt.date(2026, 12, 7), sha_prefix="s44c-sch"
        )
        await session.commit()
    report_id, _, _ = await build_ready_core(stage3_session_maker, cycle_id=cycle_id)
    settings = _settings(openrouter_api_key="k", openrouter_model="m")
    async with stage3_session_maker() as session:
        pos = await session.scalar(select(DebtPosition).limit(1))
        assert pos is not None
        position_id = int(pos.id)
        pos.comment_raw = "уточнить схему"
        await session.commit()
    async with stage3_session_maker() as session:
        job_id = await enqueue_enrichment_job(
            session, report_id=report_id, settings=settings
        )
    async with stage3_session_maker() as session:
        claim = await claim_enrichment_job(
            session, job_id=job_id, settings=settings
        )
    assert claim is not None

    class BadSchema:
        async def analyze_comment(self, **kwargs: object) -> LlmCommentFacts:
            raise OpenRouterSchemaError("invalid date")

    result = await run_claimed_enrichment(
        stage3_session_maker,
        claim=claim,
        settings=settings,
        llm_client=BadSchema(),  # type: ignore[arg-type]
    )
    assert result is not None
    async with stage3_session_maker() as session:
        job = await session.get(CommentEnrichmentJob, job_id)
        assert job is not None
        assert job.status is CommentEnrichmentJobStatus.READY
        row = await session.scalar(
            select(CommentAnalysis).where(
                CommentAnalysis.enrichment_job_id == job_id,
                CommentAnalysis.debt_position_id == position_id,
            )
        )
        assert row is not None
        assert row.analysis_status is CommentAnalysisStatus.NEEDS_REVIEW
        assert row.source is CommentAnalysisSource.UNPARSED
        assert row.parse_notes and "schema_invalid" in row.parse_notes


@pytest.mark.asyncio
async def test_frozen_snapshot_ignores_debt_position_mutation(
    stage3_session_maker: async_sessionmaker[AsyncSession],
    valid_result: ValidationResult,
) -> None:
    async with stage3_session_maker() as session:
        cycle_id = await complete_cycle(
            session, valid_result, report_date=dt.date(2026, 12, 8), sha_prefix="s44c-snap"
        )
        await session.commit()
    report_id, _, _ = await build_ready_core(stage3_session_maker, cycle_id=cycle_id)
    settings = _settings()
    position_id = await _set_only_comment(
        stage3_session_maker, comment="Оплата 08.12 на 100"
    )
    async with stage3_session_maker() as session:
        job_id = await enqueue_enrichment_job(
            session, report_id=report_id, settings=settings
        )
    async with stage3_session_maker() as session:
        job = await session.get(CommentEnrichmentJob, job_id)
        assert job is not None
        frozen_hash = str(job.enrichment_input_hash)
        frozen_comment = str(job.input_snapshot_json[0]["comment_raw"])
    async with stage3_session_maker() as session:
        pos = await session.get(DebtPosition, position_id)
        assert pos is not None
        pos.comment_raw = "МУТАЦИЯ ПОСЛЕ SNAPSHOT"
        await session.commit()
    async with stage3_session_maker() as session:
        claim = await claim_enrichment_job(
            session, job_id=job_id, settings=settings
        )
    assert claim is not None
    result = await run_claimed_enrichment(
        stage3_session_maker, claim=claim, settings=settings, llm_client=None
    )
    assert result is not None
    async with stage3_session_maker() as session:
        job = await session.get(CommentEnrichmentJob, job_id)
        assert job is not None
        assert job.enrichment_input_hash == frozen_hash
        artifact = await session.get(AuditArtifact, result.artifact_id)
        assert artifact is not None
        wb = load_workbook(BytesIO(artifact.excel_bytes))
        sheet = wb["Комментарии"]
        values = [
            str(cell.value)
            for row in sheet.iter_rows(min_row=2, values_only=False)
            for cell in row
            if cell.value is not None
        ]
        joined = " ".join(values)
        assert frozen_comment in joined
        assert "МУТАЦИЯ ПОСЛЕ SNAPSHOT" not in joined


@pytest.mark.asyncio
async def test_on_delete_restrict_comment_analysis(
    stage3_session_maker: async_sessionmaker[AsyncSession],
    valid_result: ValidationResult,
) -> None:
    async with stage3_session_maker() as session:
        cycle_id = await complete_cycle(
            session, valid_result, report_date=dt.date(2026, 12, 9), sha_prefix="s44c-fk"
        )
        await session.commit()
    report_id, _, _ = await build_ready_core(stage3_session_maker, cycle_id=cycle_id)
    settings = _settings()
    async with stage3_session_maker() as session:
        pos = await session.scalar(select(DebtPosition).limit(1))
        assert pos is not None
        position_id = int(pos.id)
        pos.comment_raw = "Оплата 09.12"
        await session.commit()
    async with stage3_session_maker() as session:
        job_id = await enqueue_enrichment_job(
            session, report_id=report_id, settings=settings
        )
    async with stage3_session_maker() as session:
        claim = await claim_enrichment_job(
            session, job_id=job_id, settings=settings
        )
    assert claim is not None
    await run_claimed_enrichment(
        stage3_session_maker, claim=claim, settings=settings, llm_client=None
    )
    async with stage3_session_maker() as session:
        with pytest.raises(IntegrityError):
            await session.execute(
                text("DELETE FROM debt_positions WHERE id = :id"),
                {"id": position_id},
            )
            await session.commit()


@pytest.mark.asyncio
async def test_recovery_automatic_only_enriched_r1(
    stage3_session_maker: async_sessionmaker[AsyncSession],
    valid_result: ValidationResult,
) -> None:
    async with stage3_session_maker() as session:
        cycle_id = await complete_cycle(
            session, valid_result, report_date=dt.date(2026, 12, 10), sha_prefix="s44c-rec"
        )
        await session.commit()
    report_id, _, _ = await build_ready_core(stage3_session_maker, cycle_id=cycle_id)
    settings = _settings()
    await _set_only_comment(stage3_session_maker, comment="Оплата 10.12")
    # Build r1 and r2
    for model in ("m-rec-1", "m-rec-2"):
        s = _settings(openrouter_model=model)
        async with stage3_session_maker() as session:
            job_id = await enqueue_enrichment_job(
                session, report_id=report_id, settings=s
            )
        async with stage3_session_maker() as session:
            claim = await claim_enrichment_job(session, job_id=job_id, settings=s)
        assert claim is not None
        await run_claimed_enrichment(
            stage3_session_maker, claim=claim, settings=s, llm_client=None
        )

    async with stage3_session_maker() as session:
        enriched = (
            await session.execute(
                select(AuditArtifact).where(
                    AuditArtifact.audit_report_id == report_id,
                    AuditArtifact.kind == AuditArtifactKind.ENRICHED,
                )
            )
        ).scalars().all()
        assert {a.revision for a in enriched} == {1, 2}
        r1 = next(a for a in enriched if a.revision == 1)
        r2 = next(a for a in enriched if a.revision == 2)
        r1_id, r2_id = int(r1.id), int(r2.id)
        # Drop automatic deliveries to exercise recovery
        await session.execute(text("DELETE FROM report_deliveries"))
        await session.commit()

    async with stage3_session_maker() as session:
        created = await recover_missing_automatic_deliveries(session)
    assert created  # CORE r1 + ENRICHED r1 delivery ids
    async with stage3_session_maker() as session:
        r2_delivery = await session.scalar(
            select(ReportDelivery).where(
                ReportDelivery.audit_artifact_id == r2_id,
                ReportDelivery.kind == ReportDeliveryKind.AUTOMATIC,
            )
        )
        assert r2_delivery is None
        r1_delivery = await session.scalar(
            select(ReportDelivery).where(
                ReportDelivery.audit_artifact_id == r1_id,
                ReportDelivery.kind == ReportDeliveryKind.AUTOMATIC,
            )
        )
        assert r1_delivery is not None
        assert r1_delivery.id in created


@pytest.mark.asyncio
async def test_enriched_document_already_sent_skips_resend(
    stage3_session_maker: async_sessionmaker[AsyncSession],
    valid_result: ValidationResult,
) -> None:
    async with stage3_session_maker() as session:
        cycle_id = await complete_cycle(
            session, valid_result, report_date=dt.date(2026, 12, 11), sha_prefix="s44c-doc"
        )
        await session.commit()
    report_id, _, _ = await build_ready_core(stage3_session_maker, cycle_id=cycle_id)
    settings = _settings()
    await _set_only_comment(stage3_session_maker, comment="Оплата 11.12")
    async with stage3_session_maker() as session:
        job_id = await enqueue_enrichment_job(
            session, report_id=report_id, settings=settings
        )
    async with stage3_session_maker() as session:
        claim = await claim_enrichment_job(
            session, job_id=job_id, settings=settings
        )
    assert claim is not None
    result = await run_claimed_enrichment(
        stage3_session_maker, claim=claim, settings=settings, llm_client=None
    )
    assert result is not None
    artifact_id = result.artifact_id

    async with stage3_session_maker() as session:
        # CORE must already be delivered so scheduler only retries ENRICHED.
        core_auto = await session.scalar(
            select(ReportDelivery)
            .join(AuditArtifact, AuditArtifact.id == ReportDelivery.audit_artifact_id)
            .where(
                AuditArtifact.kind == AuditArtifactKind.CORE,
                ReportDelivery.kind == ReportDeliveryKind.AUTOMATIC,
            )
        )
        assert core_auto is not None
        core_auto.status = ReportDeliveryStatus.DELIVERED
        core_auto.document_message_id = 1
        core_auto.delivered_at = dt.datetime.now(dt.timezone.utc)
        delivery = await session.scalar(
            select(ReportDelivery).where(
                ReportDelivery.audit_artifact_id == artifact_id,
                ReportDelivery.kind == ReportDeliveryKind.AUTOMATIC,
            )
        )
        assert delivery is not None
        delivery_id = int(delivery.id)
        await session.commit()
    async with stage3_session_maker() as session:
        dclaim = await claim_delivery(
            session, delivery_id=delivery_id, settings=settings
        )
    assert dclaim is not None
    async with stage3_session_maker() as session:
        await record_document_sent(
            session,
            delivery_id=delivery_id,
            claim_token=dclaim.claim_token,
            message_id=9999,
        )
        delivery = await session.get(ReportDelivery, delivery_id)
        assert delivery is not None
        delivery.status = ReportDeliveryStatus.FAILED
        delivery.claim_token = None
        delivery.claimed_at = None
        delivery.next_retry_at = dt.datetime.now(dt.timezone.utc) - dt.timedelta(
            seconds=1
        )
        await session.commit()

    send_document = AsyncMock()

    async def fake_send_message(**kwargs: object) -> object:
        return SimpleNamespace(message_id=1)

    scheduler = ReportScheduler(
        bot=AsyncMock(),
        session_maker=stage3_session_maker,
        settings=settings,  # type: ignore[arg-type]
        send_message=fake_send_message,
        send_document=send_document,
    )
    await scheduler.run_once()
    send_document.assert_not_awaited()
    async with stage3_session_maker() as session:
        delivery = await session.get(ReportDelivery, delivery_id)
        assert delivery is not None
        assert delivery.status == ReportDeliveryStatus.DELIVERED
        assert delivery.document_message_id == 9999


@pytest.mark.asyncio
async def test_enriched_telegram_summary_uses_job_counts(
    stage3_session_maker: async_sessionmaker[AsyncSession],
    valid_result: ValidationResult,
) -> None:
    async with stage3_session_maker() as session:
        cycle_id = await complete_cycle(
            session, valid_result, report_date=dt.date(2026, 12, 12), sha_prefix="s44c-sum"
        )
        await session.commit()
    report_id, _, _ = await build_ready_core(stage3_session_maker, cycle_id=cycle_id)
    settings = _settings()
    await _set_only_comment(stage3_session_maker, comment="Оплата 12.12 на 10")
    async with stage3_session_maker() as session:
        job_id = await enqueue_enrichment_job(
            session, report_id=report_id, settings=settings
        )
    async with stage3_session_maker() as session:
        claim = await claim_enrichment_job(
            session, job_id=job_id, settings=settings
        )
    assert claim is not None
    result = await run_claimed_enrichment(
        stage3_session_maker, claim=claim, settings=settings, llm_client=None
    )
    assert result is not None

    async with stage3_session_maker() as session:
        report = await session.get(AuditReport, report_id)
        assert report is not None
        original_summary = dict(report.summary_json or {})
        job = await session.get(CommentEnrichmentJob, job_id)
        assert job is not None
        counts = dict(job.enrichment_counts_json or {})
        auto = await session.scalar(
            select(ReportDelivery).where(
                ReportDelivery.audit_artifact_id == result.artifact_id,
                ReportDelivery.kind == ReportDeliveryKind.AUTOMATIC,
            )
        )
        assert auto is not None
        auto_id = int(auto.id)
        core = await session.scalar(
            select(AuditArtifact).where(
                AuditArtifact.audit_report_id == report_id,
                AuditArtifact.kind == AuditArtifactKind.CORE,
            )
        )
        assert core is not None
        core_delivery = await session.scalar(
            select(ReportDelivery).where(
                ReportDelivery.audit_artifact_id == core.id,
                ReportDelivery.kind == ReportDeliveryKind.AUTOMATIC,
            )
        )
        assert core_delivery is not None
        core_delivery_id = int(core_delivery.id)

    async with stage3_session_maker() as session:
        ctx = await load_delivery_send_context(session, delivery_id=auto_id)
        assert ctx is not None
        joined = "\n".join(ctx.summary_messages)
        assert "Анализ комментариев" in joined
        assert f"Всего комментариев: {counts['total']}" in joined
        assert "LLM:" in joined

    async with stage3_session_maker() as session:
        report2 = await session.get(AuditReport, report_id)
        assert report2 is not None
        assert report2.summary_json == original_summary

    async with stage3_session_maker() as session:
        manual_id = await create_manual_delivery(
            session,
            artifact_id=result.artifact_id,
            destination_chat_id=NOTIFY_CHAT_ID,
            requested_by_user_id=1,
        )

    async with stage3_session_maker() as session:
        mctx = await load_delivery_send_context(session, delivery_id=manual_id)
        assert mctx is not None
        assert "Анализ комментариев" in "\n".join(mctx.summary_messages)

    async with stage3_session_maker() as session:
        cctx = await load_delivery_send_context(session, delivery_id=core_delivery_id)
        assert cctx is not None
        assert "Анализ комментариев" not in "\n".join(cctx.summary_messages)


@pytest.mark.asyncio
async def test_redaction_before_openrouter_no_raw_pii(
    stage3_session_maker: async_sessionmaker[AsyncSession],
    valid_result: ValidationResult,
) -> None:
    async with stage3_session_maker() as session:
        cycle_id = await complete_cycle(
            session, valid_result, report_date=dt.date(2026, 12, 13), sha_prefix="s44c-red"
        )
        await session.commit()
    report_id, _, _ = await build_ready_core(stage3_session_maker, cycle_id=cycle_id)
    settings = _settings(openrouter_api_key="k", openrouter_model="m")
    pii_comment = (
        "уточнить позже, email manager@example.com phone +7(916)123-45-67 "
        "р/с 40817810099910004312"
    )
    await _set_only_comment(stage3_session_maker, comment=pii_comment)
    async with stage3_session_maker() as session:
        job_id = await enqueue_enrichment_job(
            session, report_id=report_id, settings=settings
        )
    async with stage3_session_maker() as session:
        claim = await claim_enrichment_job(
            session, job_id=job_id, settings=settings
        )
    assert claim is not None
    captured: list[dict[str, object]] = []

    class CapturingClient:
        async def analyze_comment(self, **kwargs: object) -> LlmCommentFacts:
            captured.append(dict(kwargs))
            return LlmCommentFacts(
                mentioned_date=None,
                mentioned_amount=None,
                action=None,
                reason=None,
                responsible_person=None,
                summary="ok",
                confidence="low",
                raw_json={"confidence": "low"},
            )

    await run_claimed_enrichment(
        stage3_session_maker,
        claim=claim,
        settings=settings,
        llm_client=CapturingClient(),  # type: ignore[arg-type]
    )
    assert captured
    sent = str(captured[0]["comment_raw"])
    assert "manager@example.com" not in sent
    assert "+7(916)123-45-67" not in sent
    assert "40817810099910004312" not in sent
    assert "[EMAIL]" in sent
    assert "[PHONE]" in sent
    assert "[ACCOUNT]" in sent
    assert b"PK\x03\x04" not in sent.encode("utf-8", errors="ignore")


@pytest.mark.asyncio
async def test_frozen_job_versions_ignore_settings_v2(
    stage3_session_maker: async_sessionmaker[AsyncSession],
    valid_result: ValidationResult,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    async with stage3_session_maker() as session:
        cycle_id = await complete_cycle(
            session, valid_result, report_date=dt.date(2026, 12, 14), sha_prefix="s44c-ver"
        )
        await session.commit()
    report_id, _, _ = await build_ready_core(stage3_session_maker, cycle_id=cycle_id)
    settings_v1 = _settings(
        openrouter_api_key="k",
        openrouter_model="model-v1",
        comment_prompt_version="1",
        comment_parser_version="1",
    )
    async with stage3_session_maker() as session:
        pos = await session.scalar(select(DebtPosition).limit(1))
        assert pos is not None
        pos.comment_raw = "уточнить версию"
        await session.commit()
    async with stage3_session_maker() as session:
        job_id = await enqueue_enrichment_job(
            session, report_id=report_id, settings=settings_v1
        )
        job = await session.get(CommentEnrichmentJob, job_id)
        assert job is not None
        assert job.model_name == "model-v1"
        assert job.prompt_version == "1"

    constructed: list[str] = []

    class FakeOR:
        def __init__(self, **kwargs: object) -> None:
            constructed.append(str(kwargs.get("model")))
            self._model = kwargs.get("model")

        async def analyze_comment(self, **kwargs: object) -> LlmCommentFacts:
            return LlmCommentFacts(
                mentioned_date=None,
                mentioned_amount=None,
                action=None,
                reason=None,
                responsible_person=None,
                summary="v1-run",
                confidence="low",
                raw_json={"confidence": "low"},
            )

    monkeypatch.setattr(
        "app.application.comment_enrichment_service.OpenRouterClient", FakeOR
    )
    settings_v2 = _settings(
        openrouter_api_key="k",
        openrouter_model="model-v2",
        comment_prompt_version="2",
        comment_parser_version="1",
    )
    async with stage3_session_maker() as session:
        claim = await claim_enrichment_job(
            session, job_id=job_id, settings=settings_v2
        )
    assert claim is not None
    # Do not inject llm_client — force factory path with frozen model_name
    result = await run_claimed_enrichment(
        stage3_session_maker, claim=claim, settings=settings_v2
    )
    assert result is not None
    assert constructed == ["model-v1"]
    async with stage3_session_maker() as session:
        job = await session.get(CommentEnrichmentJob, job_id)
        assert job is not None
        assert job.model_name == "model-v1"
        assert job.prompt_version == "1"
        analysis = await session.scalar(
            select(CommentAnalysis).where(CommentAnalysis.enrichment_job_id == job_id)
        )
        assert analysis is not None
        expected = compute_analysis_input_hash(
            comment_raw="уточнить версию",
            report_date=dt.date(2026, 12, 14),
            versions={
                "parser_version": "1",
                "prompt_version": "1",
                "schema_version_llm": "1",
                "redaction_version": "1",
                "model_name": "model-v1",
            },
        )
        assert analysis.analysis_input_hash == expected
        artifact = await session.get(AuditArtifact, result.artifact_id)
        assert artifact is not None
        assert artifact.prompt_version == "1"
        assert artifact.model_name == "model-v1"


@pytest.mark.asyncio
async def test_unsupported_frozen_parser_version_fails_controlled(
    stage3_session_maker: async_sessionmaker[AsyncSession],
    valid_result: ValidationResult,
) -> None:
    async with stage3_session_maker() as session:
        cycle_id = await complete_cycle(
            session, valid_result, report_date=dt.date(2026, 12, 15), sha_prefix="s44c-uns"
        )
        await session.commit()
    report_id, _, _ = await build_ready_core(stage3_session_maker, cycle_id=cycle_id)
    settings = _settings()
    async with stage3_session_maker() as session:
        pos = await session.scalar(select(DebtPosition).limit(1))
        assert pos is not None
        pos.comment_raw = "Оплата 15.12"
        await session.commit()
    async with stage3_session_maker() as session:
        job_id = await enqueue_enrichment_job(
            session, report_id=report_id, settings=settings
        )
        job = await session.get(CommentEnrichmentJob, job_id)
        assert job is not None
        job.parser_version = "99"
        await session.commit()
    async with stage3_session_maker() as session:
        claim = await claim_enrichment_job(
            session, job_id=job_id, settings=settings
        )
    assert claim is not None
    result = await run_claimed_enrichment(
        stage3_session_maker, claim=claim, settings=settings, llm_client=None
    )
    assert result is None
    async with stage3_session_maker() as session:
        job = await session.get(CommentEnrichmentJob, job_id)
        assert job is not None
        assert job.status is CommentEnrichmentJobStatus.FAILED
        assert job.last_error and "unsupported_frozen_parser_version" in job.last_error
