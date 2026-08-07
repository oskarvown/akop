"""Пустой хвост колонок справа от Q и пустые строки внизу — артефакты Excel.

См. `docs/DATA_CONTRACT.md` §4: после обрезки хвоста fingerprint по-прежнему
требует ровно 17 колонок; колонка с данными правее Q отклоняет файл.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

from app.infrastructure.excel.fingerprint import EXPECTED_COLUMN_COUNT, SHEET_NAME
from app.infrastructure.excel.reader import read_workbook
from app.infrastructure.excel.validator import validate_confirmed_template_file


def _copy_valid_xlsx_with_trailing_empty_extent(fixtures_dir: Path, dest: Path) -> Path:
    """Копия валидного fixture + искусственно расширенный max_column/max_row.

    Пишем пустую строку в далёкую ячейку (колонка Z) и несколько пустых строк
    ниже «Итого» — openpyxl увеличивает used range, но содержимого нет; ридер
    должен обрезать хвост до 17 колонок и до строки «Итого».
    """
    source = fixtures_dir / "regional_valid_basic.xlsx"
    workbook = openpyxl.load_workbook(source)
    worksheet = workbook[SHEET_NAME]
    for col_idx in range(18, 27):
        letter = get_column_letter(col_idx)
        worksheet.column_dimensions[letter].width = 12.0
    # Пустая строка расширяет max_column, но обрезается как пустое содержимое
    worksheet.cell(row=1, column=26).value = ""
    last = worksheet.max_row or 1
    for offset in range(1, 6):
        worksheet.cell(row=last + offset, column=1).value = None
    workbook.save(dest)
    workbook.close()
    return dest


def test_trailing_empty_columns_and_rows_are_trimmed(
    fixtures_dir: Path, tmp_path: Path
) -> None:
    path = _copy_valid_xlsx_with_trailing_empty_extent(
        fixtures_dir, tmp_path / "valid_with_trailing_empty.xlsx"
    )

    raw_openpyxl = openpyxl.load_workbook(path, data_only=True)
    assert (raw_openpyxl[SHEET_NAME].max_column or 0) > EXPECTED_COLUMN_COUNT
    raw_openpyxl.close()

    sheet = read_workbook(path, SHEET_NAME)
    assert sheet.n_cols == EXPECTED_COLUMN_COUNT
    assert sheet.rows[-1].values[0]
    assert str(sheet.rows[-1].values[0]).lower().startswith("итог")

    result = validate_confirmed_template_file(path)
    assert result.is_valid is True
    assert result.parsed is not None


def test_extra_column_with_data_beyond_q_is_not_trimmed_away(
    fixtures_dir: Path, tmp_path: Path
) -> None:
    """Реальные данные в колонке R — это уже не хвост, файл отклоняется."""
    path = tmp_path / "extra_data_column.xlsx"
    source = fixtures_dir / "regional_valid_basic.xlsx"
    workbook = openpyxl.load_workbook(source)
    worksheet = workbook[SHEET_NAME]
    worksheet.cell(row=12, column=18).value = "лишняя колонка с данными"
    workbook.save(path)
    workbook.close()

    sheet = read_workbook(path, SHEET_NAME)
    assert sheet.n_cols == 18

    result = validate_confirmed_template_file(path)
    assert result.is_valid is False
    joined = " | ".join(result.rejection_reasons)
    assert "17" in joined
    assert "18" in joined
