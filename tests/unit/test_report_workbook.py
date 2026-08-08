"""Unit tests for Stage 4.2 CORE workbook builder (corrective)."""
from __future__ import annotations

import datetime as dt
import time
from decimal import Decimal
from io import BytesIO

import openpyxl

from app.application.comparison_service import (
    ADDITIVE_METRICS,
    ControlEquality,
    CycleComparison,
    MatchedEntity,
    MetricDelta,
    PositionSnapshot,
    build_l1_control_equalities,
    compare_position_sets,
    compute_metric_deltas,
)
from app.application.report_workbook import (
    ALL_METRIC_HEADERS,
    CHANGES_HEADERS,
    CONTRACT_HEADERS,
    CONTROL_HEADERS,
    COUNTERPARTY_HEADERS,
    DEPARTMENT_HEADERS,
    DOCUMENT_HEADERS,
    MANAGER_GROUP_HEADERS,
    MONEY_FORMAT,
    SHEET_HEADERS,
    SHEET_ORDER,
    SUMMARY_HEADERS,
    aggregate_metric_deltas,
    build_core_excel_bytes,
    build_department_rows,
    build_manager_group_rows,
    change_class,
    sanitize_excel_text,
)


def _metrics(
    *,
    total: Decimal | None = Decimal("100"),
    extra: dict[str, Decimal | None] | None = None,
) -> dict[str, Decimal | None]:
    values = {
        "document_amount": Decimal("10"),
        "total_debt": total,
        "advance": Decimal("1"),
        "not_due": Decimal("2"),
        "overdue_1_7": Decimal("3"),
        "overdue_8_14": Decimal("4"),
        "overdue_15_21": Decimal("5"),
        "overdue_22_30": Decimal("6"),
        "overdue_over_31": Decimal("7"),
    }
    if extra:
        values.update(extra)
    return values


def _pos(
    *,
    pid: int,
    key: str,
    level: int = 1,
    total: Decimal | None = Decimal("100"),
    label: str = "Acme",
    department: str = "regional",
    manager_group_id: int = 1,
    counterparty_id: int = 1,
    parent_position_id: int | None = None,
    manager_group_name: str = "Ivanov",
    counterparty_name: str = "Acme LLC",
    metrics_extra: dict[str, Decimal | None] | None = None,
) -> PositionSnapshot:
    return PositionSnapshot(
        id=pid,
        match_key=key,
        match_key_hash="h" * 64,
        outline_level=level,
        raw_label=label,
        counterparty_id=counterparty_id,
        manager_group_id=manager_group_id,
        source_file_id=1,
        department=department,
        metrics=_metrics(total=total, extra=metrics_extra),
        credit_limit=None,
        parent_position_id=parent_position_id,
        manager_group_name=manager_group_name,
        counterparty_name=counterparty_name,
    )


def _entity(
    curr: PositionSnapshot | None,
    prev: PositionSnapshot | None,
    *,
    ambiguous: bool = False,
    collision_current_count: int | None = None,
    collision_previous_count: int | None = None,
) -> MatchedEntity:
    level = (curr or prev).outline_level  # type: ignore[union-attr]
    key = (curr or prev).match_key  # type: ignore[union-attr]
    if ambiguous:
        deltas = {
            name: MetricDelta(None, None, None, None) for name in ADDITIVE_METRICS
        }
    else:
        deltas = compute_metric_deltas(curr=curr, prev=prev)
    return MatchedEntity(
        match_key=key,
        outline_level=level,
        current=None if ambiguous else curr,
        previous=None if ambiguous else prev,
        ambiguous=ambiguous,
        collision_current_count=(
            2
            if ambiguous and collision_current_count is None
            else (
                collision_current_count
                if collision_current_count is not None
                else (1 if curr else 0)
            )
        ),
        collision_previous_count=(
            collision_previous_count
            if collision_previous_count is not None
            else (0 if ambiguous else (1 if prev else 0))
        ),
        deltas=deltas,
        overdue_profile_changed=False,
        document_bucket_transition=None,
    )


def _comparison(
    entities: list[MatchedEntity],
    *,
    current_positions: list[PositionSnapshot] | None = None,
    previous_positions: list[PositionSnapshot] | None = None,
    collisions=(),
    previous_cycle_id: int | None = 9,
) -> CycleComparison:
    if current_positions is None:
        current_positions = [
            e.current for e in entities if e.current is not None
        ]
    if previous_positions is None:
        previous_positions = [
            e.previous for e in entities if e.previous is not None
        ]
    return CycleComparison(
        current_cycle_id=10,
        current_report_date=dt.date(2026, 8, 8),
        previous_cycle_id=previous_cycle_id,
        previous_report_date=(
            dt.date(2026, 8, 1) if previous_cycle_id is not None else None
        ),
        entities=tuple(entities),
        collisions=tuple(collisions),
        control_equalities=(
            ControlEquality(
                name="company_total_debt_vs_sum_departments",
                left=Decimal("100"),
                right=Decimal("100"),
                ok=True,
            ),
        ),
        ambiguous_keys=frozenset(c.match_key for c in collisions),
        current_positions=tuple(current_positions),
        previous_positions=tuple(previous_positions),
    )


def _header_row(ws) -> list[str]:
    return [ws.cell(1, c).value for c in range(1, ws.max_column + 1) if ws.cell(1, c).value]


def test_sanitize_formula_injection() -> None:
    assert sanitize_excel_text("=CMD()") == "'=CMD()"
    assert sanitize_excel_text("+1+1") == "'+1+1"
    assert sanitize_excel_text("-1") == "'-1"
    assert sanitize_excel_text("@SUM(A1)") == "'@SUM(A1)"
    assert sanitize_excel_text("\tHIJACK") == "'\tHIJACK"
    assert sanitize_excel_text("normal") == "normal"


def test_change_class_labels() -> None:
    curr = _pos(pid=1, key="c:1", total=Decimal("120"))
    prev = _pos(pid=2, key="c:1", total=Decimal("100"))
    assert "рост" in change_class(_entity(curr, prev))
    assert "чистое снижение" in change_class(
        _entity(
            _pos(pid=1, key="c:1", total=Decimal("50")),
            _pos(pid=2, key="c:1", total=Decimal("100")),
        )
    )
    assert change_class(_entity(curr, None)) == "новая позиция"
    assert change_class(_entity(None, prev)) == "закрыта"


def test_sheet_order_and_exact_headers() -> None:
    curr = _pos(pid=1, key="c:1", total=Decimal("100"))
    prev = _pos(pid=2, key="c:1", total=Decimal("80"))
    raw, _ = build_core_excel_bytes(_comparison([_entity(curr, prev)]))
    wb = openpyxl.load_workbook(BytesIO(raw))
    assert wb.sheetnames == list(SHEET_ORDER)
    assert SUMMARY_HEADERS == SHEET_HEADERS["Сводка"]
    assert DEPARTMENT_HEADERS == SHEET_HEADERS["Отделы"]
    assert MANAGER_GROUP_HEADERS == SHEET_HEADERS["ManagerGroup"]
    assert COUNTERPARTY_HEADERS == SHEET_HEADERS["Контрагенты"]
    assert CONTRACT_HEADERS == SHEET_HEADERS["Договоры"]
    assert DOCUMENT_HEADERS == SHEET_HEADERS["Документы"]
    assert CHANGES_HEADERS == SHEET_HEADERS["Изменения"]
    assert CONTROL_HEADERS == SHEET_HEADERS["Контроль"]
    for name in SHEET_ORDER:
        assert _header_row(wb[name])[: len(SHEET_HEADERS[name])] == list(
            SHEET_HEADERS[name]
        )


def test_all_additive_metrics_have_four_columns() -> None:
    assert len(ALL_METRIC_HEADERS) == 9 * 4
    for metric in ADDITIVE_METRICS:
        for suffix in ("current", "previous", "abs_delta", "percent_delta"):
            assert f"{metric} {suffix}" in ALL_METRIC_HEADERS

    curr = _pos(pid=1, key="c:1", total=Decimal("120"))
    prev = _pos(pid=2, key="c:1", total=Decimal("100"))
    raw, _ = build_core_excel_bytes(_comparison([_entity(curr, prev)]))
    wb = openpyxl.load_workbook(BytesIO(raw))
    for sheet_name in (
        "Отделы",
        "ManagerGroup",
        "Контрагенты",
        "Договоры",
        "Документы",
        "Изменения",
    ):
        headers = _header_row(wb[sheet_name])
        for metric in ADDITIVE_METRICS:
            for suffix in ("current", "previous", "abs_delta", "percent_delta"):
                assert f"{metric} {suffix}" in headers

    summary_headers = _header_row(wb["Сводка"])
    assert summary_headers[:5] == list(SUMMARY_HEADERS)
    assert wb["Сводка"].max_row >= 1 + len(ADDITIVE_METRICS)


def test_baseline_without_previous() -> None:
    curr = _pos(pid=1, key="c:1", total=Decimal("100"))
    comparison = _comparison(
        [_entity(curr, None)],
        previous_cycle_id=None,
        previous_positions=[],
    )
    raw, _ = build_core_excel_bytes(comparison)
    wb = openpyxl.load_workbook(BytesIO(raw))
    ws = wb["Контрагенты"]
    headers = _header_row(ws)
    new_col = headers.index("new") + 1
    closed_col = headers.index("closed") + 1
    prev_col = headers.index("total_debt previous") + 1
    pct_col = headers.index("total_debt percent_delta") + 1
    assert ws.cell(2, new_col).value == "да"
    assert ws.cell(2, closed_col).value == "нет"
    assert ws.cell(2, prev_col).value is None
    assert ws.cell(2, pct_col).value == "н/д"


def test_new_and_closed_flags() -> None:
    opened = _pos(pid=1, key="c:new", total=Decimal("40"), label="NewCo")
    closed = _pos(pid=2, key="c:old", total=Decimal("55"), label="OldCo")
    comparison = _comparison(
        [_entity(opened, None), _entity(None, closed)],
        current_positions=[opened],
        previous_positions=[closed],
    )
    raw, _ = build_core_excel_bytes(comparison)
    wb = openpyxl.load_workbook(BytesIO(raw))
    ws = wb["Контрагенты"]
    headers = _header_row(ws)
    key_col = headers.index("match_key") + 1
    new_col = headers.index("new") + 1
    closed_col = headers.index("closed") + 1
    by_key = {
        ws.cell(r, key_col).value: (
            ws.cell(r, new_col).value,
            ws.cell(r, closed_col).value,
        )
        for r in range(2, ws.max_row + 1)
    }
    assert by_key["c:new"] == ("да", "нет")
    assert by_key["c:old"] == ("нет", "да")


def test_prev_zero_percent_is_nd() -> None:
    curr = _pos(pid=1, key="c:1", total=Decimal("10"))
    prev = _pos(pid=2, key="c:1", total=Decimal("0"))
    deltas = compute_metric_deltas(curr=curr, prev=prev)
    assert deltas["total_debt"].percent_delta is None
    raw, _ = build_core_excel_bytes(_comparison([_entity(curr, prev)]))
    wb = openpyxl.load_workbook(BytesIO(raw))
    ws = wb["Контрагенты"]
    headers = _header_row(ws)
    pct_col = headers.index("total_debt percent_delta") + 1
    assert ws.cell(2, pct_col).value == "н/д"


def test_l2_l3_l4_ancestry_columns() -> None:
    l1 = _pos(pid=1, key="c:1", level=1, label="Acme", counterparty_name="Acme")
    l2 = _pos(
        pid=2,
        key="c:1|2:dogovor",
        level=2,
        label="=Dogovor",
        parent_position_id=1,
    )
    l3 = _pos(
        pid=3,
        key="c:1|2:dogovor|3:object",
        level=3,
        label="+Object",
        parent_position_id=2,
    )
    l4 = _pos(
        pid=4,
        key="c:1|2:dogovor|3:object|4:doc",
        level=4,
        label="@Doc",
        parent_position_id=3,
    )
    entities = [
        _entity(l1, None),
        _entity(l2, None),
        _entity(l3, None),
        _entity(l4, None),
    ]
    comparison = _comparison(
        entities,
        current_positions=[l1, l2, l3, l4],
        previous_positions=[],
        previous_cycle_id=None,
    )
    raw, _ = build_core_excel_bytes(comparison)
    wb = openpyxl.load_workbook(BytesIO(raw))
    contracts = wb["Договоры"]
    docs = wb["Документы"]
    c_headers = _header_row(contracts)
    d_headers = _header_row(docs)
    assert contracts.cell(2, c_headers.index("Договор") + 1).value == "'=Dogovor"
    # L3 row on contracts sheet
    obj_vals = [
        contracts.cell(r, c_headers.index("Объект") + 1).value
        for r in range(2, contracts.max_row + 1)
    ]
    assert "'+Object" in obj_vals
    assert docs.cell(2, d_headers.index("Договор") + 1).value == "'=Dogovor"
    assert docs.cell(2, d_headers.index("Объект") + 1).value == "'+Object"
    assert docs.cell(2, d_headers.index("Ярлык") + 1).value == "'@Doc"


def test_real_compare_position_sets_collision_builds_workbook() -> None:
    a = _pos(
        pid=1,
        key="c:dup",
        total=Decimal("30"),
        label="=Alpha",
        manager_group_name="=Mgr",
        counterparty_name="+CP",
    )
    b = _pos(
        pid=2,
        key="c:dup",
        total=Decimal("70"),
        label="-Beta",
        manager_group_name="=Mgr",
        counterparty_name="+CP",
    )
    unique = _pos(pid=3, key="c:ok", total=Decimal("10"), label="OK")
    entities, collisions, ambiguous = compare_position_sets(
        current=[a, b, unique], previous=[]
    )
    assert "c:dup" in ambiguous
    assert collisions
    comparison = CycleComparison(
        current_cycle_id=1,
        current_report_date=dt.date(2026, 8, 8),
        previous_cycle_id=None,
        previous_report_date=None,
        entities=entities,
        collisions=collisions,
        control_equalities=(),
        ambiguous_keys=ambiguous,
        current_positions=(a, b, unique),
        previous_positions=(),
    )
    raw, digest = build_core_excel_bytes(comparison)
    assert len(digest) == 64
    wb = openpyxl.load_workbook(BytesIO(raw))

    control = wb["Контроль"]
    types = [control.cell(r, 1).value for r in range(2, control.max_row + 1)]
    assert "collision" in types
    raw_labels_col = _header_row(control).index("raw_labels") + 1
    label_cell = next(
        control.cell(r, raw_labels_col).value
        for r in range(2, control.max_row + 1)
        if control.cell(r, 1).value == "collision"
    )
    assert "'=Alpha" in str(label_cell)
    assert "'-Beta" in str(label_cell)

    changes = wb["Изменения"]
    ch_headers = _header_row(changes)
    keys = [
        changes.cell(r, ch_headers.index("match_key") + 1).value
        for r in range(2, changes.max_row + 1)
    ]
    assert "c:dup" in keys
    amb_row = next(
        r
        for r in range(2, changes.max_row + 1)
        if changes.cell(r, ch_headers.index("match_key") + 1).value == "c:dup"
    )
    assert changes.cell(amb_row, ch_headers.index("ambiguous") + 1).value == "да"
    assert changes.cell(amb_row, ch_headers.index("collision_current_count") + 1).value == 2

    # Aggregates include collision L1 amounts: 30+70+10 = 110
    company = aggregate_metric_deltas(
        current_positions=comparison.current_positions,
        previous_positions=comparison.previous_positions,
    )
    assert company["total_debt"].current == Decimal("110")
    summary = wb["Сводка"]
    metrics = {
        summary.cell(r, 1).value: summary.cell(r, 2).value
        for r in range(2, 2 + len(ADDITIVE_METRICS))
    }
    assert metrics["total_debt"] == 110


def test_freeze_autofilter_and_money_formats() -> None:
    curr = _pos(pid=1, key="c:1", total=Decimal("100"))
    prev = _pos(pid=2, key="c:1", total=Decimal("80"))
    raw, _ = build_core_excel_bytes(_comparison([_entity(curr, prev)]))
    wb = openpyxl.load_workbook(BytesIO(raw))
    for name in SHEET_ORDER:
        ws = wb[name]
        assert ws.freeze_panes == "A2"
        assert ws.auto_filter.ref is not None
        assert ws.auto_filter.ref.startswith("A1:")
    counterparties = wb["Контрагенты"]
    headers = _header_row(counterparties)
    curr_col = headers.index("total_debt current") + 1
    assert counterparties.cell(2, curr_col).number_format == MONEY_FORMAT
    assert isinstance(counterparties.cell(2, curr_col).value, (int, float, Decimal))


def test_empty_money_cells_keep_money_format_after_roundtrip() -> None:
    # Baseline without previous + one NULL metric + control equality with empty sides.
    curr = _pos(
        pid=1,
        key="c:1",
        total=Decimal("100"),
        metrics_extra={"advance": None},
    )
    comparison = _comparison(
        [_entity(curr, None)],
        previous_cycle_id=None,
        previous_positions=[],
        current_positions=[curr],
    )
    comparison = CycleComparison(
        current_cycle_id=comparison.current_cycle_id,
        current_report_date=comparison.current_report_date,
        previous_cycle_id=None,
        previous_report_date=None,
        entities=comparison.entities,
        collisions=(),
        control_equalities=(
            ControlEquality(
                name="l2_l4_disclosure_only",
                left=None,
                right=None,
                ok=True,
                diagnostic=True,
            ),
        ),
        ambiguous_keys=frozenset(),
        current_positions=(curr,),
        previous_positions=(),
    )
    raw, _ = build_core_excel_bytes(comparison)
    wb = openpyxl.load_workbook(BytesIO(raw))

    money_suffixes = ("current", "previous", "abs_delta")
    for sheet_name in (
        "Сводка",
        "Отделы",
        "ManagerGroup",
        "Контрагенты",
        "Договоры",
        "Документы",
        "Изменения",
    ):
        ws = wb[sheet_name]
        headers = _header_row(ws)
        if sheet_name == "Сводка":
            for col_name in ("current", "previous", "abs_delta"):
                col = headers.index(col_name) + 1
                for row in range(2, 2 + len(ADDITIVE_METRICS)):
                    assert ws.cell(row, col).number_format == MONEY_FORMAT
            continue
        for metric in ADDITIVE_METRICS:
            for suffix in money_suffixes:
                header = f"{metric} {suffix}"
                if header not in headers:
                    continue
                col = headers.index(header) + 1
                for row in range(2, ws.max_row + 1):
                    assert ws.cell(row, col).number_format == MONEY_FORMAT, (
                        sheet_name,
                        header,
                        row,
                    )

    # Explicit NULL/baseline spots.
    cp = wb["Контрагенты"]
    cp_headers = _header_row(cp)
    assert cp.cell(2, cp_headers.index("total_debt previous") + 1).value is None
    assert (
        cp.cell(2, cp_headers.index("total_debt previous") + 1).number_format
        == MONEY_FORMAT
    )
    assert cp.cell(2, cp_headers.index("advance current") + 1).value is None
    assert (
        cp.cell(2, cp_headers.index("advance current") + 1).number_format
        == MONEY_FORMAT
    )
    assert cp.cell(2, cp_headers.index("advance abs_delta") + 1).value is None
    assert (
        cp.cell(2, cp_headers.index("advance abs_delta") + 1).number_format
        == MONEY_FORMAT
    )

    control = wb["Контроль"]
    ctrl_headers = _header_row(control)
    left_col = ctrl_headers.index("left") + 1
    right_col = ctrl_headers.index("right") + 1
    disclosure_row = next(
        r
        for r in range(2, control.max_row + 1)
        if control.cell(r, ctrl_headers.index("имя") + 1).value
        == "l2_l4_disclosure_only"
    )
    assert control.cell(disclosure_row, left_col).value is None
    assert control.cell(disclosure_row, right_col).value is None
    assert control.cell(disclosure_row, left_col).number_format == MONEY_FORMAT
    assert control.cell(disclosure_row, right_col).number_format == MONEY_FORMAT


def test_aggregates_match_stage41_control_style_sum() -> None:
    p1 = _pos(
        pid=1,
        key="c:1",
        total=Decimal("40"),
        department="regional",
        manager_group_id=1,
        counterparty_id=1,
        metrics_extra={
            "document_amount": Decimal("11"),
            "advance": Decimal("2"),
            "not_due": Decimal("3"),
            "overdue_1_7": Decimal("4"),
            "overdue_8_14": Decimal("5"),
            "overdue_15_21": Decimal("6"),
            "overdue_22_30": Decimal("7"),
            "overdue_over_31": Decimal("8"),
        },
    )
    p2 = _pos(
        pid=2,
        key="c:2",
        total=Decimal("60"),
        department="moscow",
        manager_group_id=2,
        counterparty_id=2,
        manager_group_name="Petrov",
        counterparty_name="Beta",
        metrics_extra={
            "document_amount": Decimal("21"),
            "advance": Decimal("12"),
            "not_due": Decimal("13"),
            "overdue_1_7": Decimal("14"),
            "overdue_8_14": Decimal("15"),
            "overdue_15_21": Decimal("16"),
            "overdue_22_30": Decimal("17"),
            "overdue_over_31": Decimal("18"),
        },
    )
    # L2–L4 with large values must not affect company/department/MG aggregates.
    l2 = _pos(
        pid=3,
        key="c:1|2:x",
        level=2,
        total=Decimal("999"),
        parent_position_id=1,
        metrics_extra={m: Decimal("1000") for m in ADDITIVE_METRICS},
    )
    l3 = _pos(
        pid=4,
        key="c:1|2:x|3:y",
        level=3,
        total=Decimal("888"),
        parent_position_id=3,
        metrics_extra={m: Decimal("2000") for m in ADDITIVE_METRICS},
    )
    l4 = _pos(
        pid=5,
        key="c:1|2:x|3:y|4:z",
        level=4,
        total=Decimal("777"),
        parent_position_id=4,
        metrics_extra={m: Decimal("3000") for m in ADDITIVE_METRICS},
    )
    positions = [p1, p2, l2, l3, l4]
    controls = build_l1_control_equalities(positions)
    control_by_name = {c.name: c for c in controls}

    entities = [
        _entity(p1, None),
        _entity(p2, None),
        _entity(l2, None),
        _entity(l3, None),
        _entity(l4, None),
    ]
    comparison = _comparison(
        entities,
        current_positions=positions,
        previous_positions=[],
        previous_cycle_id=None,
    )
    raw, _ = build_core_excel_bytes(comparison)
    wb = openpyxl.load_workbook(BytesIO(raw))

    # Company: workbook Сводка current == Control company rollup left.
    summary = wb["Сводка"]
    summary_current = {
        summary.cell(r, 1).value: summary.cell(r, 2).value
        for r in range(2, 2 + len(ADDITIVE_METRICS))
    }
    for metric in ADDITIVE_METRICS:
        expected = control_by_name[f"company_{metric}_vs_sum_departments"].left
        assert Decimal(str(summary_current[metric])) == expected

    # Department sheet currents match Control department left sides.
    dept_sheet = wb["Отделы"]
    dept_headers = _header_row(dept_sheet)
    dept_rows = {
        dept_sheet.cell(r, dept_headers.index("department_key") + 1).value: r
        for r in range(2, dept_sheet.max_row + 1)
    }
    for dept, row in dept_rows.items():
        for metric in ADDITIVE_METRICS:
            expected = control_by_name[
                f"department_{dept}_{metric}_vs_sum_manager_groups"
            ].left
            col = dept_headers.index(f"{metric} current") + 1
            assert Decimal(str(dept_sheet.cell(row, col).value)) == expected

    # ManagerGroup sheet currents match Control MG left sides.
    mg_sheet = wb["ManagerGroup"]
    mg_headers = _header_row(mg_sheet)
    mg_rows = {
        int(mg_sheet.cell(r, mg_headers.index("manager_group_id") + 1).value): r
        for r in range(2, mg_sheet.max_row + 1)
    }
    for mg_id, row in mg_rows.items():
        for metric in ADDITIVE_METRICS:
            expected = control_by_name[
                f"manager_group_{mg_id}_{metric}_vs_sum_counterparties"
            ].left
            col = mg_headers.index(f"{metric} current") + 1
            assert Decimal(str(mg_sheet.cell(row, col).value)) == expected

    # Helper aggregates agree with the same Control totals.
    company = aggregate_metric_deltas(
        current_positions=positions, previous_positions=[]
    )
    for metric in ADDITIVE_METRICS:
        assert (
            company[metric].current
            == control_by_name[f"company_{metric}_vs_sum_departments"].left
        )
    for row in build_department_rows(comparison):
        for metric in ADDITIVE_METRICS:
            assert (
                row.deltas[metric].current
                == control_by_name[
                    f"department_{row.department_key}_{metric}_vs_sum_manager_groups"
                ].left
            )
    for row in build_manager_group_rows(comparison):
        assert row.manager_group_id is not None
        for metric in ADDITIVE_METRICS:
            assert (
                row.deltas[metric].current
                == control_by_name[
                    f"manager_group_{row.manager_group_id}_{metric}"
                    "_vs_sum_counterparties"
                ].left
            )


def test_rebuild_same_bytes_and_sha() -> None:
    curr = _pos(pid=1, key="c:1", total=Decimal("150"))
    prev = _pos(pid=2, key="c:1", total=Decimal("100"))
    comparison = _comparison([_entity(curr, prev)])
    bytes1, sha1 = build_core_excel_bytes(comparison)
    time.sleep(1.05)
    bytes2, sha2 = build_core_excel_bytes(comparison)
    assert sha1 == sha2
    assert bytes1 == bytes2


def test_formula_injection_on_user_fields_and_collision_labels() -> None:
    a = _pos(
        pid=1,
        key="c:1",
        label="Doc",
        manager_group_name="=MgrEvil",
        counterparty_name="+CpEvil",
    )
    b = _pos(
        pid=2,
        key="c:1",
        label="-Dup",
        manager_group_name="=MgrEvil",
        counterparty_name="+CpEvil",
    )
    entities, collisions, ambiguous = compare_position_sets(
        current=[a, b], previous=[]
    )
    comparison = CycleComparison(
        current_cycle_id=1,
        current_report_date=dt.date(2026, 8, 8),
        previous_cycle_id=None,
        previous_report_date=None,
        entities=entities,
        collisions=collisions,
        control_equalities=(),
        ambiguous_keys=ambiguous,
        current_positions=(a, b),
        previous_positions=(),
    )
    raw, _ = build_core_excel_bytes(comparison)
    wb = openpyxl.load_workbook(BytesIO(raw))
    changes = wb["Изменения"]
    headers = _header_row(changes)
    mg = changes.cell(2, headers.index("ManagerGroup") + 1).value
    cp = changes.cell(2, headers.index("Контрагент") + 1).value
    labels = changes.cell(2, headers.index("raw_labels") + 1).value
    assert str(mg).startswith("'=")
    assert str(cp).startswith("'+")
    assert "'-Dup" in str(labels) or str(labels).startswith("'-")
