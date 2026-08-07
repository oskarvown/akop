"""Генератор обезличенных fixture-файлов отдела «Региональный» (Stage 2).

Все данные (названия менеджеров/контрагентов/договоров, суммы) — полностью
синтетические, не связаны с реальными файлами из `private_inputs/` (см.
`docs/ASSUMPTIONS.md` §3: реальные названия/значения не переносятся в
коммитируемые fixtures). Структура (17 физических колонок, многострочная
шапка, outline levels 0–4, строка «Итого») воспроизводит подтверждённый
fingerprint из `docs/DATA_CONTRACT.md` §3.

Модуль используется как библиотека (`build_regional_xls`/`build_regional_xlsx`)
и как скрипт, генерирующий статические fixture-файлы в `tests/fixtures/regional/`
(см. `if __name__ == "__main__"` внизу).
"""
from __future__ import annotations

from dataclasses import dataclass, field
from decimal import Decimal
from pathlib import Path

FIXTURES_DIR = Path(__file__).parent / "regional"

HIERARCHY_HEADER_LABELS = ("Менеджер", "Партнер", "Договор", "Объект расчетов", "Расчетный документ")

BUCKET_HEADER_ROW = {
    10: "Не просрочено",
    11: "От 1 до 7 дней",
    12: "От 8 до 14 дней",
    13: "От 15 до 21 дней",
    14: "От 22 до 30 дней",
    15: "Свыше 31 дней",
}
MONEY_HEADER_ROW = {
    5: "Отсрочка платежа",
    6: "Сумма кредита",
    7: "Сумма документа",
    8: "Долг",
    9: "Аванс",
    10: "Задолженность",
    11: "Задолженность",
    12: "Задолженность",
    13: "Задолженность",
    14: "Задолженность",
    15: "Задолженность",
}


@dataclass(frozen=True)
class Money:
    payment_deferral_days: object = None  # int | None | "invalid-marker"
    credit_limit: Decimal | None = None
    document_amount: Decimal | None = None
    total_debt: Decimal | None = None
    advance: Decimal | None = None
    not_due: Decimal | None = None
    overdue_1_7: Decimal | None = None
    overdue_8_14: Decimal | None = None
    overdue_15_21: Decimal | None = None
    overdue_22_30: Decimal | None = None
    overdue_over_31: Decimal | None = None
    comment: str | None = None

    def as_row_values(self, label: str, level: int) -> list:
        row = [None] * 17
        row[0] = label
        if level == 4:
            row[3] = "УПД-000123"
        row[5] = self.payment_deferral_days
        row[6] = _f(self.credit_limit)
        row[7] = _f(self.document_amount)
        row[8] = _f(self.total_debt)
        row[9] = _f(self.advance)
        row[10] = _f(self.not_due)
        row[11] = _f(self.overdue_1_7)
        row[12] = _f(self.overdue_8_14)
        row[13] = _f(self.overdue_15_21)
        row[14] = _f(self.overdue_22_30)
        row[15] = _f(self.overdue_over_31)
        row[16] = self.comment
        return row


def _f(value: Decimal | None) -> float | None:
    return float(value) if value is not None else None


@dataclass(frozen=True)
class DocumentSpec:
    label: str
    money: Money = field(default_factory=Money)


@dataclass(frozen=True)
class ObjectSpec:
    label: str
    money: Money = field(default_factory=Money)
    documents: tuple[DocumentSpec, ...] = ()


@dataclass(frozen=True)
class ContractSpec:
    label: str
    money: Money = field(default_factory=Money)
    objects: tuple[ObjectSpec, ...] = ()


@dataclass(frozen=True)
class CounterpartySpec:
    label: str
    money: Money = field(default_factory=Money)
    contracts: tuple[ContractSpec, ...] = ()


@dataclass(frozen=True)
class ManagerGroupSpec:
    label: str
    counterparties: tuple[CounterpartySpec, ...]


@dataclass(frozen=True)
class WorkbookSpec:
    report_date: str  # DD.MM.YYYY
    manager_groups: tuple[ManagerGroupSpec, ...]
    grand_total: Money
    hidden_columns: tuple[int, ...] = ()
    narrow_columns: tuple[int, ...] = ()


def _iter_data_rows(spec: WorkbookSpec):
    """Yields (outline_level, row_values) в порядке физических строк файла."""
    for mg in spec.manager_groups:
        yield 0, Money().as_row_values(mg.label, 0)
        for cp in mg.counterparties:
            yield 1, cp.money.as_row_values(cp.label, 1)
            for contract in cp.contracts:
                yield 2, contract.money.as_row_values(contract.label, 2)
                for obj in contract.objects:
                    yield 3, obj.money.as_row_values(obj.label, 3)
                    for doc in obj.documents:
                        yield 4, doc.money.as_row_values(doc.label, 4)
    yield 0, spec.grand_total.as_row_values("Итого", 0)


def _header_rows(spec: WorkbookSpec) -> list[tuple[int, list]]:
    param_row = [None] * 17
    param_row[2] = f"Дата отчета: {spec.report_date}"
    rows: list[tuple[int, list]] = [
        (0, param_row),
        (1, _labelled_row("Параметры:")),
        (1, _labelled_row("")),
        (1, _labelled_row("Отбор:")),
        (0, _labelled_row("")),
    ]
    bucket_row = [None] * 17
    bucket_row[0] = "Менеджер"
    for col, text in BUCKET_HEADER_ROW.items():
        bucket_row[col] = text
    money_row = [None] * 17
    money_row[0] = "Партнер"
    for col, text in MONEY_HEADER_ROW.items():
        money_row[col] = text
    rows.append((0, bucket_row))
    rows.append((0, money_row))
    rows.append((0, _labelled_row("Договор")))
    rows.append((0, _labelled_row("Объект расчетов")))
    doc_header = _labelled_row("Расчетный документ")
    doc_header[3] = "УПД"
    rows.append((0, doc_header))
    return rows


def _labelled_row(label: str) -> list:
    row = [None] * 17
    row[0] = label
    return row


def _all_rows(spec: WorkbookSpec) -> list[tuple[int, list]]:
    return _header_rows(spec) + list(_iter_data_rows(spec))


def build_regional_xls(spec: WorkbookSpec, path: Path) -> None:
    import xlwt

    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet("TDSheet")
    for col in range(17):
        sheet.col(col).width = 4000
    for col in spec.hidden_columns:
        sheet.col(col).hidden = True
    for col in spec.narrow_columns:
        sheet.col(col).width = 42

    for row_idx, (level, values) in enumerate(_all_rows(spec)):
        sheet.row(row_idx).level = level
        for col_idx, value in enumerate(values):
            if value is None:
                continue
            sheet.write(row_idx, col_idx, value)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(str(path))


def build_regional_xlsx(spec: WorkbookSpec, path: Path) -> None:
    import openpyxl
    from openpyxl.utils import get_column_letter

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "TDSheet"

    for col in spec.hidden_columns:
        letter = get_column_letter(col + 1)
        sheet.column_dimensions[letter].hidden = True
    for col in spec.narrow_columns:
        letter = get_column_letter(col + 1)
        sheet.column_dimensions[letter].width = 1

    for row_idx, (level, values) in enumerate(_all_rows(spec), start=1):
        sheet.row_dimensions[row_idx].outlineLevel = level
        for col_idx, value in enumerate(values, start=1):
            if value is None:
                continue
            sheet.cell(row=row_idx, column=col_idx, value=value)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(str(path))


def build_invalid_missing_columns_xlsx(report_date: str, path: Path) -> None:
    """12-колоночный негативный fixture (аналог `invalid_2026-07-15_missing_columns`).

    Отсутствуют: «Отсрочка платежа», «Сумма кредита», «Сумма документа»,
    «Не просрочено», корзина «От 15 до 21 дня» — см. `docs/DATA_CONTRACT.md` §8.
    """
    import openpyxl

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "TDSheet"

    rows: list[list] = [
        [None] * 12,
        _row12("Параметры:", extra={2: f"Дата отчета: {report_date}"}),
        [None] * 12,
        _row12("Отбор:"),
        [None] * 12,
    ]
    bucket_row = [None] * 12
    bucket_row[0] = "Менеджер"
    bucket_row[7] = "От 1 до 7 дней"
    bucket_row[8] = "От 8 до 14 дней"
    bucket_row[9] = "От 22 до 30 дней"
    bucket_row[10] = "Свыше 31 дней"
    money_row = [None] * 12
    money_row[0] = "Партнер"
    money_row[5] = "Долг"
    money_row[6] = "Аванс"
    money_row[7] = "Задолженность"
    money_row[8] = "Задолженность"
    money_row[9] = "Задолженность"
    money_row[10] = "Задолженность"
    rows.append(bucket_row)
    rows.append(money_row)
    rows.append(_row12("Договор"))
    rows.append(_row12("Объект расчетов"))
    doc_row = _row12("Расчетный документ")
    doc_row[4] = "УПД"
    rows.append(doc_row)

    data_row = [None] * 12
    data_row[0] = "Синтетический Менеджер"
    data_row[5] = 33312790.25
    data_row[6] = 6084472.64
    data_row[7] = 4593554.29
    data_row[8] = 139485.04
    data_row[9] = 216927.65
    data_row[10] = 2911463.76
    rows.append(data_row)

    for row_idx, values in enumerate(rows, start=1):
        for col_idx, value in enumerate(values, start=1):
            if value is None:
                continue
            sheet.cell(row=row_idx, column=col_idx, value=value)

    # Явно фиксируем физическую ширину листа в 12 колонок (индексы 0–11), как в
    # реальном `invalid_2026-07-15_missing_columns.xlsx` (§8) — иначе openpyxl
    # может определить `max_column` как 11, если последняя колонка нигде не заполнена.
    sheet.cell(row=1, column=12, value="")
    sheet.column_dimensions["L"].width = 8.0

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(str(path))


def _row12(label: str, extra: dict[int, str] | None = None) -> list:
    row = [None] * 12
    row[0] = label
    if extra:
        for col, value in extra.items():
            row[col] = value
    return row
